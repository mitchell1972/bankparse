#!/usr/bin/env python3
"""
Download an Apify dataset via REST API and save as JSON.
Then process into deduplicated CSV + Excel.

Usage:
    python scripts/download_apify_dataset.py --dataset-id y2o7FuCNBU9oeKGu3
"""

import argparse
import csv
import json
import re
import requests
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


FIELDNAMES = [
    "name", "address", "city", "state", "zip", "county",
    "phone", "email", "website", "facebook",
    "category", "categories", "rating", "reviews",
    "google_maps_url", "place_id",
]


def infer_county(search_string: str) -> str:
    m = re.search(r"in\s+(.+?)\s+County", search_string or "")
    return m.group(1) if m else ""


def extract_firm(place: dict) -> dict:
    emails = place.get("emails") or []
    facebooks = place.get("facebooks") or []
    return {
        "name": place.get("title", ""),
        "address": place.get("address", ""),
        "city": place.get("city", ""),
        "state": place.get("state", ""),
        "zip": place.get("postalCode", ""),
        "county": infer_county(place.get("searchString", "")),
        "phone": place.get("phone", ""),
        "email": emails[0] if emails else "",
        "website": place.get("website", ""),
        "facebook": facebooks[0] if facebooks else "",
        "category": place.get("categoryName", ""),
        "categories": ", ".join(place.get("categories") or []),
        "rating": place.get("totalScore", ""),
        "reviews": place.get("reviewsCount", ""),
        "google_maps_url": place.get("url", ""),
        "place_id": place.get("placeId", ""),
    }


def deduplicate(firms: list[dict]) -> list[dict]:
    seen = set()
    unique = []
    for f in firms:
        pid = f["place_id"]
        if pid and pid not in seen:
            seen.add(pid)
            unique.append(f)
    return unique


def download_dataset(dataset_id: str) -> list[dict]:
    """Download full dataset from Apify REST API (no token needed for own datasets)."""
    url = f"https://api.apify.com/v2/datasets/{dataset_id}/items"
    params = {"format": "json", "clean": "true"}

    print(f"Downloading dataset {dataset_id}...", flush=True)
    resp = requests.get(url, params=params, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    print(f"Downloaded {len(data)} raw items", flush=True)
    return data


def write_csv(firms: list[dict], output_path: Path):
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(firms)
    print(f"CSV saved: {output_path} ({len(firms)} firms)", flush=True)


def write_excel(firms: list[dict], output_path: Path):
    if not HAS_OPENPYXL:
        print("openpyxl not installed, skipping Excel output", flush=True)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "All Accounting Firms"

    header_map = {
        "name": "Firm Name", "address": "Address", "city": "City",
        "state": "State", "zip": "ZIP", "county": "County",
        "phone": "Phone", "email": "Email", "website": "Website",
        "facebook": "Facebook", "category": "Primary Category",
        "categories": "All Categories", "rating": "Rating",
        "reviews": "Reviews", "google_maps_url": "Google Maps",
        "place_id": "Place ID",
    }

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    col_widths = {
        "name": 35, "address": 45, "city": 18, "state": 12, "zip": 10,
        "county": 15, "phone": 18, "email": 30, "website": 35,
        "facebook": 35, "category": 25, "categories": 40, "rating": 8,
        "reviews": 10, "google_maps_url": 30, "place_id": 30,
    }

    def write_sheet(ws, data_firms):
        for col_idx, field in enumerate(FIELDNAMES, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_map.get(field, field))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_idx, firm in enumerate(data_firms, 2):
            for col_idx, field in enumerate(FIELDNAMES, 1):
                val = firm.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        for col_idx, field in enumerate(FIELDNAMES, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 20)
        ws.freeze_panes = "A2"
        if data_firms:
            ws.auto_filter.ref = ws.dimensions

    # Sheet 1: All firms
    write_sheet(ws, firms)

    # Sheet 2: Firms with emails only
    email_firms = [f for f in firms if f.get("email")]
    ws_email = wb.create_sheet("Firms with Emails")
    write_sheet(ws_email, email_firms)

    # Sheet 3: Summary by County
    ws2 = wb.create_sheet("Summary by County")
    county_counts = {}
    email_counts = {}
    for firm in firms:
        county = firm.get("county") or "Unknown"
        county_counts[county] = county_counts.get(county, 0) + 1
        if firm.get("email"):
            email_counts[county] = email_counts.get(county, 0) + 1

    summary_headers = ["County", "Firms Found", "With Email", "Email %"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, (county, count) in enumerate(sorted(county_counts.items()), 2):
        emails = email_counts.get(county, 0)
        pct = f"{emails/count*100:.0f}%" if count > 0 else "0%"
        ws2.cell(row=i, column=1, value=county)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=emails)
        ws2.cell(row=i, column=4, value=pct)
        if i % 2 == 0:
            for c in range(1, 5):
                ws2.cell(row=i, column=c).fill = alt_fill

    total_row = len(county_counts) + 2
    total_emails = sum(email_counts.values())
    total_firms = len(firms)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=total_firms).font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=total_emails).font = Font(bold=True)
    ws2.cell(row=total_row, column=4, value=f"{total_emails/total_firms*100:.0f}%" if total_firms else "0%").font = Font(bold=True)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 12
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)
    print(f"Excel saved: {output_path}", flush=True)
    print(f"  Sheet 1: {total_firms} firms (all)", flush=True)
    print(f"  Sheet 2: {len(email_firms)} firms WITH emails", flush=True)
    print(f"  Sheet 3: Summary by County", flush=True)


def main():
    parser = argparse.ArgumentParser(description="Download Apify dataset and convert to CSV/Excel")
    parser.add_argument("--dataset-id", required=True, help="Apify dataset ID")
    parser.add_argument("--output", default="accountants_alabama", help="Output base name")
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent.parent

    # Download
    raw_data = download_dataset(args.dataset_id)

    # Save raw JSON
    raw_path = project_root / f"{args.output}_raw.json"
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(raw_data, f, indent=2)
    print(f"Raw JSON saved: {raw_path}", flush=True)

    # Extract and deduplicate
    firms = [extract_firm(place) for place in raw_data]
    firms = deduplicate(firms)
    firms.sort(key=lambda x: (x["county"], x["name"]))

    print(f"\n{len(raw_data)} raw → {len(firms)} unique firms (deduped by Place ID)", flush=True)

    email_count = sum(1 for f in firms if f.get("email"))
    print(f"Firms with email: {email_count}/{len(firms)} ({email_count/len(firms)*100:.0f}%)", flush=True)

    # Write outputs
    csv_path = project_root / f"{args.output}.csv"
    xlsx_path = project_root / f"{args.output}.xlsx"

    write_csv(firms, csv_path)
    write_excel(firms, xlsx_path)


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
Convert accountants CSV to a formatted Excel workbook.

Usage:
    python scripts/csv_to_excel.py                                          # defaults
    python scripts/csv_to_excel.py --input accountants_alabama.csv          # custom input
    python scripts/csv_to_excel.py --input accountants_alabama.csv --output accountants_alabama.xlsx
"""

import argparse
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def csv_to_excel(input_path: Path, output_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounting Firms"

    # Read CSV
    with open(input_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        print("CSV is empty.")
        return

    headers = rows[0]
    data = rows[1:]

    # Friendly header names
    header_map = {
        "name": "Firm Name",
        "address": "Address",
        "phone": "Phone",
        "intl_phone": "Intl Phone",
        "website": "Website",
        "google_maps_url": "Google Maps",
        "business_status": "Status",
        "types": "Categories",
        "place_id": "Place ID",
        "county": "County",
        "state": "State",
    }

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_map.get(h, h))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto-fit column widths (with max)
    col_widths = {
        "name": 35,
        "address": 45,
        "phone": 18,
        "intl_phone": 20,
        "website": 40,
        "google_maps_url": 30,
        "business_status": 15,
        "types": 40,
        "place_id": 30,
        "county": 15,
        "state": 12,
    }
    for col_idx, h in enumerate(headers, 1):
        width = col_widths.get(h, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary by County")
    county_col = headers.index("county") if "county" in headers else None
    if county_col is not None:
        county_counts = {}
        for row in data:
            county = row[county_col]
            county_counts[county] = county_counts.get(county, 0) + 1

        ws2.cell(row=1, column=1, value="County").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=1).alignment = header_align
        ws2.cell(row=1, column=2, value="Firms Found").font = header_font
        ws2.cell(row=1, column=2).fill = header_fill
        ws2.cell(row=1, column=2).alignment = header_align

        for i, (county, count) in enumerate(sorted(county_counts.items()), 2):
            ws2.cell(row=i, column=1, value=county)
            ws2.cell(row=i, column=2, value=count)
            if i % 2 == 0:
                ws2.cell(row=i, column=1).fill = alt_fill
                ws2.cell(row=i, column=2).fill = alt_fill

        total_row = len(county_counts) + 2
        ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws2.cell(row=total_row, column=2, value=len(data)).font = Font(bold=True)

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 15
        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    print(f"  Sheet 1: {len(data)} firms across {len(county_counts) if county_col is not None else '?'} counties")
    print(f"  Sheet 2: Summary by County")


def main():
    parser = argparse.ArgumentParser(description="Convert accountants CSV to Excel")
    parser.add_argument("--input", default="accountants_alabama.csv", help="Input CSV")
    parser.add_argument("--output", default=None, help="Output Excel path (defaults to same name .xlsx)")
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent.parent
    input_path = project_root / args.input
    output_path = project_root / (args.output or args.input.replace(".csv", ".xlsx"))

    if not input_path.exists():
        print(f"ERROR: {input_path} not found")
        return

    csv_to_excel(input_path, output_path)


if __name__ == "__main__":
    main()

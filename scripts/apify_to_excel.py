#!/usr/bin/env python3
"""
Convert Apify Google Maps Scraper JSON output to CSV + formatted Excel.

Usage:
    python scripts/apify_to_excel.py --input apify_results.json
    python scripts/apify_to_excel.py --input apify_results.json --output accountants_alabama
"""

import argparse
import csv
import json
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


FIELDNAMES = [
    "name", "address", "city", "state", "zip", "county",
    "phone", "email", "website", "facebook",
    "category", "categories", "rating", "reviews",
    "google_maps_url", "place_id",
]


def infer_county(search_string: str) -> str:
    """Extract county name from the Apify searchString field."""
    m = re.search(r"in\s+(.+?)\s+County", search_string or "")
    return m.group(1) if m else ""


def extract_firm(place: dict) -> dict:
    """Extract structured data from an Apify Places result."""
    emails = place.get("emails") or []
    phones = place.get("phones") or []
    phones_uncertain = place.get("phonesUncertain") or []
    facebooks = place.get("facebooks") or []

    return {
        "name": place.get("title", ""),
        "address": place.get("address", ""),
        "city": place.get("city", ""),
        "state": place.get("state", ""),
        "zip": place.get("postalCode", ""),
        "county": infer_county(place.get("searchString", "")),
        "phone": place.get("phone", "") or (phones[0] if phones else ""),
        "email": emails[0] if emails else "",
        "website": place.get("website", ""),
        "facebook": facebooks[0] if facebooks else "",
        "category": place.get("categoryName", ""),
        "categories": ", ".join(place.get("categories") or []),
        "rating": place.get("totalScore", ""),
        "reviews": place.get("reviewsCount", ""),
        "google_maps_url": place.get("url", ""),
        "place_id": place.get("placeId", ""),
    }


def deduplicate(firms: list[dict]) -> list[dict]:
    """Deduplicate by place_id, keeping the first occurrence."""
    seen = set()
    unique = []
    for f in firms:
        pid = f["place_id"]
        if pid and pid not in seen:
            seen.add(pid)
            unique.append(f)
    return unique


def write_csv(firms: list[dict], output_path: Path):
    """Write firms to CSV."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(firms)
    print(f"CSV saved: {output_path} ({len(firms)} firms)")


def write_excel(firms: list[dict], output_path: Path):
    """Write firms to formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Accounting Firms"

    # Friendly header names
    header_map = {
        "name": "Firm Name",
        "address": "Address",
        "city": "City",
        "state": "State",
        "zip": "ZIP",
        "county": "County",
        "phone": "Phone",
        "email": "Email",
        "website": "Website",
        "facebook": "Facebook",
        "category": "Primary Category",
        "categories": "All Categories",
        "rating": "Rating",
        "reviews": "Reviews",
        "google_maps_url": "Google Maps",
        "place_id": "Place ID",
    }

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Write headers
    for col_idx, field in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_map.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, firm in enumerate(firms, 2):
        for col_idx, field in enumerate(FIELDNAMES, 1):
            val = firm.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    col_widths = {
        "name": 35, "address": 45, "city": 18, "state": 12, "zip": 10,
        "county": 15, "phone": 18, "email": 30, "website": 35,
        "facebook": 35, "category": 25, "categories": 40, "rating": 8,
        "reviews": 10, "google_maps_url": 30, "place_id": 30,
    }
    for col_idx, field in enumerate(FIELDNAMES, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # --- Firms with Emails sheet ---
    ws_email = wb.create_sheet("Firms with Emails")
    email_firms = [f for f in firms if f.get("email")]

    for col_idx, field in enumerate(FIELDNAMES, 1):
        cell = ws_email.cell(row=1, column=col_idx, value=header_map.get(field, field))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, firm in enumerate(email_firms, 2):
        for col_idx, field in enumerate(FIELDNAMES, 1):
            val = firm.get(field, "")
            cell = ws_email.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx, field in enumerate(FIELDNAMES, 1):
        ws_email.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(field, 20)
    ws_email.freeze_panes = "A2"
    ws_email.auto_filter.ref = ws_email.dimensions

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary by County")
    county_counts = {}
    email_counts = {}
    for firm in firms:
        county = firm.get("county", "Unknown")
        county_counts[county] = county_counts.get(county, 0) + 1
        if firm.get("email"):
            email_counts[county] = email_counts.get(county, 0) + 1

    summary_headers = ["County", "Firms Found", "With Email", "Email %"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, (county, count) in enumerate(sorted(county_counts.items()), 2):
        emails = email_counts.get(county, 0)
        pct = f"{emails/count*100:.0f}%" if count > 0 else "0%"
        ws2.cell(row=i, column=1, value=county)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=emails)
        ws2.cell(row=i, column=4, value=pct)
        if i % 2 == 0:
            for c in range(1, 5):
                ws2.cell(row=i, column=c).fill = alt_fill

    total_row = len(county_counts) + 2
    total_emails = sum(email_counts.values())
    total_firms = len(firms)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=2, value=total_firms).font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=total_emails).font = Font(bold=True)
    ws2.cell(row=total_row, column=4, value=f"{total_emails/total_firms*100:.0f}%" if total_firms else "0%").font = Font(bold=True)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 12
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    print(f"  Sheet 1: {total_firms} firms across {len(county_counts)} counties")
    print(f"  Sheet 2: {len(email_firms)} firms WITH emails")
    print(f"  Sheet 3: Summary by County ({total_emails} firms with emails)")


def main():
    parser = argparse.ArgumentParser(description="Convert Apify results to CSV + Excel")
    parser.add_argument("--input", required=True, help="Input JSON file from Apify")
    parser.add_argument("--output", default="accountants_alabama", help="Output base name (without extension)")
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent.parent
    input_path = project_root / args.input

    with open(input_path, "r", encoding="utf-8") as f:
        raw_data = json.load(f)

    # Extract and deduplicate
    firms = [extract_firm(place) for place in raw_data]
    firms = deduplicate(firms)
    firms.sort(key=lambda x: (x["county"], x["name"]))

    print(f"Loaded {len(raw_data)} raw results → {len(firms)} unique firms")

    csv_path = project_root / f"{args.output}.csv"
    xlsx_path = project_root / f"{args.output}.xlsx"

    write_csv(firms, csv_path)
    write_excel(firms, xlsx_path)


if __name__ == "__main__":
    main()

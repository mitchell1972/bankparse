"""
Excel Exporter
Generates clean, formatted XLSX spreadsheets from parsed bank statement data.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Styling constants
HEADER_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CREDIT_FONT = Font(name="Calibri", color="1A7A2E", size=10)
DEBIT_FONT = Font(name="Calibri", color="C0392B", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B4F72")
SUMMARY_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D5D8DC"),
)

# Currency format strings keyed by ISO code
_CURRENCY_FORMATS = {
    "GBP": '£#,##0.00',
    "USD": '$#,##0.00',
    "EUR": '€#,##0.00',
    "CAD": 'C$#,##0.00',
    "AUD": 'A$#,##0.00',
    "JPY": '¥#,##0',
    "INR": '₹#,##0.00',
    "CHF": 'CHF#,##0.00',
    "NZD": 'NZ$#,##0.00',
    "SEK": 'kr#,##0.00',
    "NOK": 'kr#,##0.00',
    "DKK": 'kr#,##0.00',
    "ZAR": 'R#,##0.00',
    "SGD": 'S$#,##0.00',
    "HKD": 'HK$#,##0.00',
}
_DEFAULT_FMT = '#,##0.00'  # no currency symbol when unknown


def _currency_fmt(metadata: dict | None = None) -> str:
    """Return the Excel number-format string for the parsed currency.

    Returns a plain number format when currency can't be determined.
    """
    cur = (metadata or {}).get("currency", "").upper()
    return _CURRENCY_FORMATS.get(cur, _DEFAULT_FMT)


def export_to_xlsx(data: dict, output_path: str) -> str:
    """
    Export parsed bank statement data to a formatted XLSX file.

    Args:
        data: Dict with these keys:
            - 'transactions' — list of {date, description, type, debit, credit,
              amount, balance}. Each row may ALSO include 'hmrc_category',
              'hmrc_confidence', 'hmrc_source' keys; if present the exporter
              renders three extra columns and a third "HMRC Summary" sheet.
            - 'summary' — optional precomputed totals dict. When empty the
              exporter computes totals from `transactions` (fixes the
              cumulative-export totals=0 bug).
            - 'metadata' — bank_name, currency.
            - 'hmrc_summary' — optional dict {income: {...}, expenses: {...},
              flagged_for_review: [...], excluded: [...], business_type, period}.
              When provided, renders the HMRC Summary sheet.
        output_path: Path for the output XLSX file

    Returns:
        Path to the created file
    """
    wb = Workbook()

    # ── Transactions Sheet ──
    ws = wb.active
    ws.title = "Transactions"

    transactions = data.get("transactions", [])
    has_hmrc = bool(data.get("hmrc_summary")) or any(
        ("hmrc_category" in tx) for tx in transactions
    )

    title_span = "A1:J1" if has_hmrc else "A1:G1"

    # Title row
    ws.merge_cells(title_span)
    ws["A1"] = "Bank Statement - Transaction Report"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # Summary section — backfill totals from transactions if caller passed an
    # empty summary (cumulative-export path used to leave this {} and the
    # numbers showed as zero on every download).
    summary = data.get("summary") or {}
    metadata = data.get("metadata", {})
    if transactions and not summary.get("total_transactions"):
        _credits = sum(float(t.get("credit") or 0) for t in transactions)
        _debits = sum(float(t.get("debit") or 0) for t in transactions)
        summary = {
            "total_transactions": len(transactions),
            "total_credits": round(_credits, 2),
            "total_debits": round(_debits, 2),
            "net": round(_credits - _debits, 2),
        }

    ws["A3"] = "Total Transactions:"
    ws["B3"] = summary.get("total_transactions", 0)
    ws["A3"].font = SUMMARY_FONT

    ws["A4"] = "Total Credits:"
    ws["B4"] = summary.get("total_credits", 0)
    ws["B4"].number_format = _currency_fmt(metadata)
    ws["A4"].font = SUMMARY_FONT
    ws["B4"].font = CREDIT_FONT

    ws["A5"] = "Total Debits:"
    ws["B5"] = summary.get("total_debits", 0)
    ws["B5"].number_format = _currency_fmt(metadata)
    ws["A5"].font = SUMMARY_FONT
    ws["B5"].font = DEBIT_FONT

    ws["A6"] = "Net:"
    ws["B6"] = summary.get("net", 0)
    ws["B6"].number_format = _currency_fmt(metadata)
    ws["A6"].font = SUMMARY_FONT

    # Headers — extra three columns when the rows carry HMRC data.
    headers = ["Date", "Description", "Type", "Debit", "Credit", "Amount", "Balance"]
    if has_hmrc:
        headers += ["HMRC Category", "Confidence", "Source"]
    header_row = 8
    last_col = len(headers)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[header_row].height = 25

    # Data rows
    for i, tx in enumerate(transactions):
        row = header_row + 1 + i

        # Date
        ws.cell(row=row, column=1, value=tx["date"]).font = NORMAL_FONT

        # Description
        ws.cell(row=row, column=2, value=tx["description"]).font = NORMAL_FONT

        # Type
        ws.cell(row=row, column=3, value=tx.get("type", "")).font = NORMAL_FONT

        # Debit
        debit_cell = ws.cell(row=row, column=4)
        if tx.get("debit"):
            debit_cell.value = tx["debit"]
            debit_cell.number_format = _currency_fmt(metadata)
            debit_cell.font = DEBIT_FONT
        else:
            debit_cell.font = NORMAL_FONT

        # Credit
        credit_cell = ws.cell(row=row, column=5)
        if tx.get("credit"):
            credit_cell.value = tx["credit"]
            credit_cell.number_format = _currency_fmt(metadata)
            credit_cell.font = CREDIT_FONT
        else:
            credit_cell.font = NORMAL_FONT

        # Amount
        amount_cell = ws.cell(row=row, column=6, value=tx["amount"])
        amount_cell.number_format = _currency_fmt(metadata)
        amount_cell.font = CREDIT_FONT if tx["amount"] >= 0 else DEBIT_FONT

        # Balance
        balance_cell = ws.cell(row=row, column=7)
        if tx.get("balance") is not None:
            balance_cell.value = tx["balance"]
            balance_cell.number_format = _currency_fmt(metadata)
        balance_cell.font = NORMAL_FONT

        # HMRC columns (only when present on the row)
        if has_hmrc:
            ws.cell(row=row, column=8, value=tx.get("hmrc_category", "")).font = NORMAL_FONT
            conf = tx.get("hmrc_confidence")
            if conf is not None:
                conf_cell = ws.cell(row=row, column=9, value=float(conf))
                conf_cell.number_format = "0%"
                conf_cell.font = NORMAL_FONT
            else:
                ws.cell(row=row, column=9).font = NORMAL_FONT
            ws.cell(row=row, column=10, value=tx.get("hmrc_source", "")).font = NORMAL_FONT

        # Cell borders across every column we used
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Column widths
    col_widths = [12, 45, 12, 14, 14, 14, 14]
    if has_hmrc:
        col_widths += [22, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze panes (freeze header row)
    ws.freeze_panes = f"A{header_row + 1}"

    # Auto-filter
    if transactions:
        last_row = header_row + len(transactions)
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(last_col)}{last_row}"

    # ── Summary Sheet ──
    ws2 = wb.create_sheet("Summary")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Statement Summary"
    ws2["A1"].font = TITLE_FONT

    summary_data = [
        ("Total Transactions", summary.get("total_transactions", 0)),
        ("Total Money In (Credits)", summary.get("total_credits", 0)),
        ("Total Money Out (Debits)", summary.get("total_debits", 0)),
        ("Net Change", summary.get("net", 0)),
    ]

    for i, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=i, column=1, value=label).font = SUMMARY_FONT
        val_cell = ws2.cell(row=i, column=2, value=value)
        if isinstance(value, (int, float)) and i > 3:
            val_cell.number_format = _currency_fmt(metadata)
        val_cell.font = SUMMARY_FONT

    # Monthly breakdown if enough data
    if transactions:
        monthly = {}
        for tx in transactions:
            month_key = tx["date"][:7]  # YYYY-MM
            if month_key not in monthly:
                monthly[month_key] = {"credits": 0, "debits": 0, "count": 0}
            if tx["amount"] > 0:
                monthly[month_key]["credits"] += tx["amount"]
            else:
                monthly[month_key]["debits"] += tx["amount"]
            monthly[month_key]["count"] += 1

        if monthly:
            row_offset = len(summary_data) + 5
            ws2.cell(row=row_offset, column=1, value="Monthly Breakdown").font = TITLE_FONT

            month_headers = ["Month", "Transactions", "Credits", "Debits", "Net"]
            for col, h in enumerate(month_headers, 1):
                cell = ws2.cell(row=row_offset + 1, column=col, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

            for i, (month, vals) in enumerate(sorted(monthly.items())):
                r = row_offset + 2 + i
                ws2.cell(row=r, column=1, value=month)
                ws2.cell(row=r, column=2, value=vals["count"])
                ws2.cell(row=r, column=3, value=round(vals["credits"], 2)).number_format = _currency_fmt(metadata)
                ws2.cell(row=r, column=4, value=round(vals["debits"], 2)).number_format = _currency_fmt(metadata)
                net = round(vals["credits"] + vals["debits"], 2)
                ws2.cell(row=r, column=5, value=net).number_format = _currency_fmt(metadata)

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 18

    # ── HMRC Summary sheet ──
    # Only added when the caller passed a `hmrc_summary` dict produced by
    # `hmrc.services.mapping.aggregate_self_employment` / `_property`.
    hmrc_summary = data.get("hmrc_summary")
    if hmrc_summary:
        _add_hmrc_summary_sheet(wb, hmrc_summary, metadata)

    # Save
    wb.save(output_path)
    return output_path


# Human-friendly labels per HMRC category code. Mirrors the dropdowns in
# templates/index.html so the spreadsheet matches what users see in the UI.
#
# Keys are the *canonical* HMRC category constants from
# `hmrc.schemas.categories` — never hardcode the string here; if HMRC
# renames a field, only the schema module changes.
from hmrc.schemas import categories as _cats  # noqa: E402

_HMRC_SE_LABELS = {
    _cats.SE_INCOME: "Turnover (income)",
    _cats.SE_OTHER_INCOME: "Other income",
    _cats.SE_EXPENSE_COST_OF_GOODS: "Cost of goods bought",
    _cats.SE_EXPENSE_CIS: "CIS subcontractor payments",
    _cats.SE_EXPENSE_STAFF: "Staff costs / wages",
    _cats.SE_EXPENSE_TRAVEL: "Travel / fuel / parking",
    _cats.SE_EXPENSE_PREMISES: "Premises (rent / utilities)",
    _cats.SE_EXPENSE_REPAIRS: "Repairs & maintenance",
    _cats.SE_EXPENSE_ADMIN: "Admin / office / software",
    _cats.SE_EXPENSE_ADVERTISING: "Advertising / marketing",
    _cats.SE_EXPENSE_ENTERTAINMENT: "Business entertainment",
    _cats.SE_EXPENSE_INTEREST: "Interest on borrowing",
    _cats.SE_EXPENSE_FINANCIAL: "Bank / financial charges",
    _cats.SE_EXPENSE_BAD_DEBT: "Bad debt written off",
    _cats.SE_EXPENSE_PROFESSIONAL: "Professional / legal fees",
    _cats.SE_EXPENSE_DEPRECIATION: "Depreciation",
    _cats.SE_EXPENSE_OTHER: "Other expense",
    _cats.EXCLUDE_OWNER_TRANSFER: "Owner transfer (excluded)",
}
_HMRC_PROP_LABELS = {
    _cats.PROP_INCOME_RENT: "Rent income",
    _cats.PROP_INCOME_PREMIUMS: "Premiums of lease grant",
    _cats.PROP_INCOME_OTHER: "Other property income",
    _cats.PROP_EXPENSE_PREMISES: "Premises running costs",
    _cats.PROP_EXPENSE_REPAIRS: "Repairs & maintenance",
    _cats.PROP_EXPENSE_FINANCIAL: "Financial costs (commercial)",
    _cats.PROP_EXPENSE_PROFESSIONAL: "Professional fees",
    _cats.PROP_EXPENSE_SERVICES: "Cost of services",
    _cats.PROP_EXPENSE_TRAVEL: "Travel",
    _cats.PROP_EXPENSE_OTHER: "Other expense",
    _cats.PROP_EXPENSE_RESIDENTIAL_FINANCIAL: "Residential mortgage interest (restricted)",
    _cats.EXCLUDE_OWNER_TRANSFER: "Owner transfer (excluded)",
}


def _add_hmrc_summary_sheet(wb, hmrc: dict, metadata: dict) -> None:
    """Build the 'HMRC Summary' sheet — the draft quarterly P&L.

    This is what an accountant or HMRC inspector will care about most.
    Shows period, income by category, expenses by category, net profit,
    a rough basic-rate tax estimate, and counts of flagged + excluded rows.
    """
    biz_type = hmrc.get("business_type", "se")
    labels = _HMRC_PROP_LABELS if biz_type == "property" else _HMRC_SE_LABELS
    biz_label = "UK Property (landlord)" if biz_type == "property" else "Self-Employment (sole trader)"

    ws = wb.create_sheet("HMRC Summary")
    ws.merge_cells("A1:C1")
    ws["A1"] = "HMRC Quarterly Submission Draft"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 28

    period = hmrc.get("period") or {}
    ws["A3"] = "Business type:"
    ws["B3"] = biz_label
    ws["A3"].font = SUMMARY_FONT
    ws["A4"] = "Period:"
    if period.get("start") and period.get("end"):
        ws["B4"] = f"{period['start']} to {period['end']}"
    else:
        ws["B4"] = "—"
    ws["A4"].font = SUMMARY_FONT

    income = hmrc.get("income") or {}
    expenses = hmrc.get("expenses") or {}
    flagged = hmrc.get("flagged_for_review") or []
    excluded = hmrc.get("excluded") or []

    total_income = round(sum(income.values()), 2)
    total_expenses = round(sum(expenses.values()), 2)
    net_profit = round(total_income - total_expenses, 2)

    fmt = _currency_fmt(metadata)

    row = 6
    ws.cell(row=row, column=1, value="INCOME").font = TITLE_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    if not income:
        ws.cell(row=row, column=1, value="(no income categorised)").font = NORMAL_FONT
        row += 1
    for cat, val in sorted(income.items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=1, value=labels.get(cat, cat)).font = NORMAL_FONT
        v_cell = ws.cell(row=row, column=2, value=val)
        v_cell.number_format = fmt
        v_cell.font = CREDIT_FONT
        row += 1
    total_i_cell = ws.cell(row=row, column=1, value="Total income")
    total_i_cell.font = SUMMARY_FONT
    v_cell = ws.cell(row=row, column=2, value=total_income)
    v_cell.number_format = fmt
    v_cell.font = SUMMARY_FONT
    row += 2

    ws.cell(row=row, column=1, value="EXPENSES").font = TITLE_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1
    if not expenses:
        ws.cell(row=row, column=1, value="(no expenses categorised)").font = NORMAL_FONT
        row += 1
    for cat, val in sorted(expenses.items(), key=lambda kv: -kv[1]):
        ws.cell(row=row, column=1, value=labels.get(cat, cat)).font = NORMAL_FONT
        v_cell = ws.cell(row=row, column=2, value=val)
        v_cell.number_format = fmt
        v_cell.font = DEBIT_FONT
        row += 1
    total_e_cell = ws.cell(row=row, column=1, value="Total expenses")
    total_e_cell.font = SUMMARY_FONT
    v_cell = ws.cell(row=row, column=2, value=total_expenses)
    v_cell.number_format = fmt
    v_cell.font = SUMMARY_FONT
    row += 2

    ws.cell(row=row, column=1, value="NET PROFIT").font = TITLE_FONT
    v_cell = ws.cell(row=row, column=2, value=net_profit)
    v_cell.number_format = fmt
    v_cell.font = TITLE_FONT
    row += 1

    # Rough basic-rate (20%) tax estimate. ITSA is more nuanced (personal
    # allowance, NICs, higher-rate bands) but this gives users a ballpark
    # number to plan against; the real number comes from HMRC's
    # Individual Calculations API once we submit.
    rough_tax = round(max(net_profit, 0) * 0.20, 2)
    ws.cell(row=row, column=1, value="Estimated tax (rough, basic rate 20%)").font = NORMAL_FONT
    v_cell = ws.cell(row=row, column=2, value=rough_tax)
    v_cell.number_format = fmt
    v_cell.font = NORMAL_FONT
    row += 2

    notes = []
    if flagged:
        notes.append(f"{len(flagged)} transaction{'s' if len(flagged) != 1 else ''} flagged for review (low confidence).")
    if excluded:
        notes.append(f"{len(excluded)} owner transfer{'s' if len(excluded) != 1 else ''} excluded from totals.")
    notes.append(
        "This sheet is a draft. Final tax figure comes from HMRC's "
        "Individual Calculations API after submission."
    )
    for note in notes:
        ws.cell(row=row, column=1, value=note).font = NORMAL_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18


def export_receipt_to_xlsx(data: dict, output_path: str) -> str:
    """
    Export parsed receipt data to a formatted XLSX file.

    Args:
        data: Dict with 'items', 'totals', and 'metadata' keys
        output_path: Path for the output XLSX file

    Returns:
        Path to the created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt Items"

    metadata = data.get("metadata", {})
    totals = data.get("totals", {})
    items = data.get("items", [])

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = f"Receipt — {metadata.get('store_name', 'Unknown Store')}"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 30

    # Receipt info
    ws["A3"] = "Store:"
    ws["B3"] = metadata.get("store_name", "")
    ws["A3"].font = SUMMARY_FONT
    ws["B3"].font = NORMAL_FONT

    ws["A4"] = "Date:"
    ws["B4"] = metadata.get("date", "")
    ws["A4"].font = SUMMARY_FONT
    ws["B4"].font = NORMAL_FONT

    ws["A5"] = "Items:"
    ws["B5"] = metadata.get("item_count", len(items))
    ws["A5"].font = SUMMARY_FONT
    ws["B5"].font = NORMAL_FONT

    # Item headers
    headers = ["#", "Item", "Qty", "Unit Price", "Total"]
    header_row = 7

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[header_row].height = 25

    # Item rows
    for i, item in enumerate(items):
        row = header_row + 1 + i

        ws.cell(row=row, column=1, value=i + 1).font = NORMAL_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=row, column=2, value=item["description"]).font = NORMAL_FONT

        ws.cell(row=row, column=3, value=item["quantity"]).font = NORMAL_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

        unit_cell = ws.cell(row=row, column=4, value=item["unit_price"])
        unit_cell.number_format = _currency_fmt(metadata)
        unit_cell.font = NORMAL_FONT

        total_cell = ws.cell(row=row, column=5, value=item["total_price"])
        total_cell.number_format = _currency_fmt(metadata)
        total_cell.font = NORMAL_FONT

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Totals section
    totals_start = header_row + len(items) + 2

    TOTALS_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

    total_rows = []
    if "subtotal" in totals:
        total_rows.append(("Subtotal", totals["subtotal"]))
    if "discount" in totals:
        total_rows.append(("Discount", totals["discount"]))
    if "tax" in totals:
        total_rows.append(("VAT / Tax", totals["tax"]))
    if "total" in totals:
        total_rows.append(("TOTAL", totals["total"]))
    if "payment" in totals:
        total_rows.append(("Payment", totals["payment"]))
    if "change" in totals:
        total_rows.append(("Change", totals["change"]))

    for i, (label, value) in enumerate(total_rows):
        r = totals_start + i
        label_cell = ws.cell(row=r, column=4, value=label)
        label_cell.font = SUMMARY_FONT
        label_cell.alignment = Alignment(horizontal="right")
        label_cell.fill = TOTALS_FILL

        val_cell = ws.cell(row=r, column=5, value=value)
        val_cell.number_format = _currency_fmt(metadata)
        val_cell.fill = TOTALS_FILL
        if label == "TOTAL":
            val_cell.font = Font(name="Calibri", bold=True, size=12, color="1B4F72")
        else:
            val_cell.font = SUMMARY_FONT

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # Freeze header
    ws.freeze_panes = f"A{header_row + 1}"

    # Auto-filter
    if items:
        last_row = header_row + len(items)
        ws.auto_filter.ref = f"A{header_row}:E{last_row}"

    wb.save(output_path)
    return output_path


def export_bulk_receipts_to_xlsx(bulk_result: dict, output_path: str) -> str:
    """
    Export combined bulk receipt data to a formatted XLSX file with two sheets.

    Sheet 1 — "All Expenses": All items from all receipts grouped by store.
    Sheet 2 — "Summary": One row per receipt with totals.

    Args:
        bulk_result: Dict from parse_receipts_bulk with 'receipts', 'combined_items',
                     'grand_total', 'receipt_count', 'total_items'
        output_path: Path for the output XLSX file

    Returns:
        Path to the created file
    """
    wb = Workbook()

    # Attempt to detect currency from the first receipt that reports one
    receipts_data = bulk_result.get("receipts", [])
    metadata = bulk_result.get("metadata", {})
    for r in receipts_data:
        if not isinstance(r, dict):
            continue
        rm = r.get("metadata")
        if rm and isinstance(rm, dict) and rm.get("currency"):
            metadata = rm
            break

    # Styling
    PURPLE_FILL = PatternFill(start_color="8E44AD", end_color="8E44AD", fill_type="solid")
    PURPLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ALT_ROW_FILL = PatternFill(start_color="F5EEF8", end_color="F5EEF8", fill_type="solid")
    GRAND_TOTAL_FILL = PatternFill(start_color="D7BDE2", end_color="D7BDE2", fill_type="solid")
    GRAND_TOTAL_FONT = Font(name="Calibri", bold=True, size=12, color="4A235A")

    # -- Sheet 1: All Expenses --
    ws1 = wb.active
    ws1.title = "All Expenses"

    # Title
    ws1.merge_cells("A1:F1")
    ws1["A1"] = f"Combined Receipt Expenses — {bulk_result['receipt_count']} receipts"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color="8E44AD")
    ws1.row_dimensions[1].height = 30

    # Headers
    headers = ["Store", "Date", "Item", "Qty", "Unit Price", "Total"]
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=header)
        cell.font = PURPLE_FONT
        cell.fill = PURPLE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[header_row].height = 25

    # Data rows — items grouped by store then date
    combined_items = bulk_result.get("combined_items", [])
    sorted_items = sorted(combined_items, key=lambda x: (x.get("store", ""), x.get("date") or ""))

    for i, item in enumerate(sorted_items):
        row = header_row + 1 + i
        use_alt = i % 2 == 1

        store_cell = ws1.cell(row=row, column=1, value=item.get("store", ""))
        store_cell.font = NORMAL_FONT
        if use_alt:
            store_cell.fill = ALT_ROW_FILL

        date_cell = ws1.cell(row=row, column=2, value=item.get("date", ""))
        date_cell.font = NORMAL_FONT
        if use_alt:
            date_cell.fill = ALT_ROW_FILL

        desc_cell = ws1.cell(row=row, column=3, value=item.get("description", ""))
        desc_cell.font = NORMAL_FONT
        if use_alt:
            desc_cell.fill = ALT_ROW_FILL

        qty_cell = ws1.cell(row=row, column=4, value=item.get("quantity", 1))
        qty_cell.font = NORMAL_FONT
        qty_cell.alignment = Alignment(horizontal="center")
        if use_alt:
            qty_cell.fill = ALT_ROW_FILL

        unit_cell = ws1.cell(row=row, column=5, value=item.get("unit_price", 0))
        unit_cell.number_format = _currency_fmt(metadata)
        unit_cell.font = NORMAL_FONT
        if use_alt:
            unit_cell.fill = ALT_ROW_FILL

        total_cell = ws1.cell(row=row, column=6, value=item.get("total_price", 0))
        total_cell.number_format = _currency_fmt(metadata)
        total_cell.font = NORMAL_FONT
        if use_alt:
            total_cell.fill = ALT_ROW_FILL

        for c in range(1, 7):
            ws1.cell(row=row, column=c).border = THIN_BORDER

    # Grand total row
    if sorted_items:
        grand_row = header_row + len(sorted_items) + 2
        label_cell = ws1.cell(row=grand_row, column=5, value="GRAND TOTAL")
        label_cell.font = GRAND_TOTAL_FONT
        label_cell.fill = GRAND_TOTAL_FILL
        label_cell.alignment = Alignment(horizontal="right")

        total_val_cell = ws1.cell(row=grand_row, column=6, value=bulk_result.get("grand_total", 0))
        total_val_cell.number_format = _currency_fmt(metadata)
        total_val_cell.font = GRAND_TOTAL_FONT
        total_val_cell.fill = GRAND_TOTAL_FILL

    # Column widths
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 40
    ws1.column_dimensions["D"].width = 8
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 14

    # Freeze panes
    ws1.freeze_panes = f"A{header_row + 1}"

    # Auto-filter
    if sorted_items:
        last_row = header_row + len(sorted_items)
        ws1.auto_filter.ref = f"A{header_row}:F{last_row}"

    # -- Sheet 2: Summary --
    ws2 = wb.create_sheet("Summary")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Receipt Summary"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="8E44AD")
    ws2.row_dimensions[1].height = 30

    # Summary headers
    summary_headers = ["Store", "Date", "Items Count", "Receipt Total"]
    summary_header_row = 3
    for col, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=summary_header_row, column=col, value=header)
        cell.font = PURPLE_FONT
        cell.fill = PURPLE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[summary_header_row].height = 25

    # One row per receipt
    receipts = [r for r in bulk_result.get("receipts", []) if isinstance(r, dict)]
    for i, receipt in enumerate(receipts):
        row = summary_header_row + 1 + i
        use_alt = i % 2 == 1

        store_cell = ws2.cell(row=row, column=1, value=receipt.get("store_name", "Unknown"))
        store_cell.font = NORMAL_FONT
        if use_alt:
            store_cell.fill = ALT_ROW_FILL

        date_cell = ws2.cell(row=row, column=2, value=receipt.get("date", ""))
        date_cell.font = NORMAL_FONT
        if use_alt:
            date_cell.fill = ALT_ROW_FILL

        count_cell = ws2.cell(row=row, column=3, value=len(receipt.get("items", [])))
        count_cell.font = NORMAL_FONT
        count_cell.alignment = Alignment(horizontal="center")
        if use_alt:
            count_cell.fill = ALT_ROW_FILL

        total_cell = ws2.cell(row=row, column=4, value=receipt.get("total", 0))
        total_cell.number_format = _currency_fmt(metadata)
        total_cell.font = NORMAL_FONT
        if use_alt:
            total_cell.fill = ALT_ROW_FILL

        for c in range(1, 5):
            ws2.cell(row=row, column=c).border = THIN_BORDER

    # Grand total at bottom of summary
    if receipts:
        grand_row = summary_header_row + len(receipts) + 2
        label_cell = ws2.cell(row=grand_row, column=3, value="GRAND TOTAL")
        label_cell.font = GRAND_TOTAL_FONT
        label_cell.fill = GRAND_TOTAL_FILL
        label_cell.alignment = Alignment(horizontal="right")

        total_val_cell = ws2.cell(row=grand_row, column=4, value=bulk_result.get("grand_total", 0))
        total_val_cell.number_format = _currency_fmt(metadata)
        total_val_cell.font = GRAND_TOTAL_FONT
        total_val_cell.fill = GRAND_TOTAL_FILL

    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 16

    wb.save(output_path)
    return output_path


def export_bulk_statements_to_xlsx(bulk_result: dict, output_path: str) -> str:
    """
    Export combined bulk statement data to a formatted XLSX file.

    Sheet 1 — "All Transactions": All transactions from all statements.
    Sheet 2 — "Summary": One row per statement with totals.
    """
    wb = Workbook()

    # Attempt to detect currency from the first statement that reports one
    stmts_data = bulk_result.get("statements", [])
    metadata = bulk_result.get("metadata", {})
    for s in stmts_data:
        if not isinstance(s, dict):
            continue
        sm = s.get("metadata")
        if sm and isinstance(sm, dict) and sm.get("currency"):
            metadata = sm
            break

    PRIMARY_FILL = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
    PRIMARY_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    CREDIT_FONT = Font(name="Calibri", color="27AE60", size=11)
    DEBIT_FONT = Font(name="Calibri", color="E74C3C", size=11)
    ALT_FILL = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
    TOTAL_FONT_STYLE = Font(name="Calibri", bold=True, size=12, color="1B4F72")

    ws1 = wb.active
    ws1.title = "All Transactions"

    count = bulk_result.get("summary", {}).get("total_transactions", 0)
    ws1.merge_cells("A1:G1")
    ws1["A1"] = f"Combined Bank Statements — {bulk_result.get('statement_count', 0)} statements, {count} transactions"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    ws1.row_dimensions[1].height = 30

    headers = ["Source", "Date", "Description", "Type", "Amount", "Balance", "Bank"]
    header_row = 3
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=header)
        cell.font = PRIMARY_FONT
        cell.fill = PRIMARY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[header_row].height = 25

    txs = bulk_result.get("all_transactions", [])
    for i, tx in enumerate(txs):
        row = header_row + 1 + i
        use_alt = i % 2 == 1

        ws1.cell(row=row, column=1, value=tx.get("source", "")).font = NORMAL_FONT
        ws1.cell(row=row, column=2, value=tx.get("date", "")).font = NORMAL_FONT
        ws1.cell(row=row, column=3, value=tx.get("description", "")).font = NORMAL_FONT
        ws1.cell(row=row, column=4, value=tx.get("type", "")).font = NORMAL_FONT

        amt = tx.get("amount", 0)
        amt_cell = ws1.cell(row=row, column=5, value=amt)
        amt_cell.number_format = _currency_fmt(metadata)
        amt_cell.font = CREDIT_FONT if amt >= 0 else DEBIT_FONT

        bal = tx.get("balance")
        bal_cell = ws1.cell(row=row, column=6, value=bal if bal is not None else "")
        if bal is not None:
            bal_cell.number_format = _currency_fmt(metadata)
        bal_cell.font = NORMAL_FONT

        ws1.cell(row=row, column=7, value=tx.get("bank", "")).font = NORMAL_FONT

        if use_alt:
            for c in range(1, 8):
                ws1.cell(row=row, column=c).fill = ALT_FILL
        for c in range(1, 8):
            ws1.cell(row=row, column=c).border = THIN_BORDER

    if txs:
        summary = bulk_result.get("summary", {})
        sr = header_row + len(txs) + 2
        for label, key, color in [("CREDITS", "total_credits", "27AE60"), ("DEBITS", "total_debits", "E74C3C"), ("NET", "net", "1B4F72")]:
            ws1.cell(row=sr, column=3, value=label).font = TOTAL_FONT_STYLE
            ws1.cell(row=sr, column=3).fill = TOTAL_FILL
            v_cell = ws1.cell(row=sr, column=5, value=summary.get(key, 0))
            v_cell.number_format = _currency_fmt(metadata)
            v_cell.font = Font(name="Calibri", bold=True, size=12, color=color)
            v_cell.fill = TOTAL_FILL
            sr += 1

    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 14
    ws1.column_dimensions["C"].width = 45
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 14
    ws1.column_dimensions["G"].width = 18
    ws1.freeze_panes = f"A{header_row + 1}"
    if txs:
        ws1.auto_filter.ref = f"A{header_row}:G{header_row + len(txs)}"

    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:E1")
    ws2["A1"] = "Statement Summary"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    ws2.row_dimensions[1].height = 30

    s_headers = ["File", "Bank", "Transactions", "Credits", "Debits"]
    s_row = 3
    for col, h in enumerate(s_headers, 1):
        cell = ws2.cell(row=s_row, column=col, value=h)
        cell.font = PRIMARY_FONT
        cell.fill = PRIMARY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    statements = [s for s in bulk_result.get("statements", []) if isinstance(s, dict)]
    for i, stmt in enumerate(statements):
        row = s_row + 1 + i
        s = stmt.get("summary", {})
        ws2.cell(row=row, column=1, value=stmt.get("source", "")).font = NORMAL_FONT
        ws2.cell(row=row, column=2, value=stmt.get("bank_name", "")).font = NORMAL_FONT
        ws2.cell(row=row, column=3, value=stmt.get("transaction_count", 0)).font = NORMAL_FONT
        cr = ws2.cell(row=row, column=4, value=s.get("total_credits", 0))
        cr.number_format = _currency_fmt(metadata)
        cr.font = CREDIT_FONT
        dr = ws2.cell(row=row, column=5, value=s.get("total_debits", 0))
        dr.number_format = _currency_fmt(metadata)
        dr.font = DEBIT_FONT
        if i % 2 == 1:
            for c in range(1, 6):
                ws2.cell(row=row, column=c).fill = ALT_FILL
        for c in range(1, 6):
            ws2.cell(row=row, column=c).border = THIN_BORDER

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16

    wb.save(output_path)
    return output_path

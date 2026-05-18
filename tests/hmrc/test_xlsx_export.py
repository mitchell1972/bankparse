"""Tests for HMRC-annotated XLSX export."""

import os
import sys
import tempfile

import pytest
from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))


def _basic_rows():
    return [
        {"date": "2025-12-22", "description": "DD TV LICENCE MBP",
         "type": "debit", "debit": 14.95, "credit": None, "amount": -14.95, "balance": None,
         "hmrc_category": "adminCosts", "hmrc_confidence": 0.80, "hmrc_source": "rule"},
        {"date": "2025-12-22", "description": "CR STRIPE PAYOUT",
         "type": "credit", "debit": None, "credit": 1500.0, "amount": 1500.0, "balance": None,
         "hmrc_category": "turnover", "hmrc_confidence": 0.85, "hmrc_source": "rule"},
        {"date": "2025-12-22", "description": "MIPERMIT LTD CHIPPENHAM",
         "type": "debit", "debit": 2.60, "credit": None, "amount": -2.60, "balance": None,
         "hmrc_category": "travelCosts", "hmrc_confidence": 0.90, "hmrc_source": "rule"},
    ]


def _summary():
    return {
        "business_type": "se",
        "period": {"start": "2025-12-22", "end": "2025-12-22"},
        "income": {"turnover": 1500.0},
        "expenses": {"adminCosts": 14.95, "travelCosts": 2.60},
        "flagged_for_review": [],
        "excluded": [],
    }


def test_xlsx_includes_hmrc_columns_when_data_present():
    from parsers.xlsx_exporter import export_to_xlsx
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        export_to_xlsx(
            {"transactions": _basic_rows(), "summary": {}, "metadata": {"currency": "GBP"},
             "hmrc_summary": _summary()},
            tf.name,
        )
        wb = load_workbook(tf.name)
    assert "Transactions" in wb.sheetnames
    assert "HMRC Summary" in wb.sheetnames
    ws = wb["Transactions"]
    headers = [c.value for c in ws[8]]  # header row is row 8
    assert "HMRC Category" in headers
    assert "Confidence" in headers
    assert "Source" in headers
    # First data row should have the expected categorisation
    cat_col = headers.index("HMRC Category") + 1
    src_col = headers.index("Source") + 1
    assert ws.cell(row=9, column=cat_col).value == "adminCosts"
    assert ws.cell(row=9, column=src_col).value == "rule"


def test_xlsx_omits_hmrc_columns_when_not_requested():
    """Backward compat: existing single-statement downloads keep the 7-column layout."""
    from parsers.xlsx_exporter import export_to_xlsx
    rows = [
        {"date": "2025-12-22", "description": "Acme Ltd", "type": "debit",
         "debit": 50.0, "credit": None, "amount": -50.0, "balance": None},
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        export_to_xlsx({"transactions": rows, "summary": {}, "metadata": {}}, tf.name)
        wb = load_workbook(tf.name)
    assert "HMRC Summary" not in wb.sheetnames
    ws = wb["Transactions"]
    headers = [c.value for c in ws[8]]
    assert headers == ["Date", "Description", "Type", "Debit", "Credit", "Amount", "Balance"]


def test_xlsx_summary_totals_computed_when_caller_passes_empty():
    """Regression for the cumulative-export bug where Total Transactions / Net
    showed as 0 because callers passed summary={}. The exporter must now
    backfill totals from the transaction rows."""
    from parsers.xlsx_exporter import export_to_xlsx
    rows = [
        {"date": "2025-12-22", "description": "X", "type": "debit",
         "debit": 14.95, "credit": None, "amount": -14.95, "balance": None},
        {"date": "2025-12-22", "description": "Y", "type": "credit",
         "debit": None, "credit": 1500.0, "amount": 1500.0, "balance": None},
    ]
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        export_to_xlsx({"transactions": rows, "summary": {}, "metadata": {}}, tf.name)
        wb = load_workbook(tf.name)
    ws = wb["Transactions"]
    assert ws["B3"].value == 2          # Total Transactions
    assert ws["B4"].value == 1500.0     # Total Credits
    assert ws["B5"].value == 14.95      # Total Debits
    assert ws["B6"].value == 1485.05    # Net


def test_hmrc_summary_sheet_renders_totals_and_estimated_tax():
    from parsers.xlsx_exporter import export_to_xlsx
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        export_to_xlsx(
            {"transactions": _basic_rows(), "summary": {}, "metadata": {"currency": "GBP"},
             "hmrc_summary": _summary()},
            tf.name,
        )
        wb = load_workbook(tf.name)
    ws = wb["HMRC Summary"]
    # Scan all cells for the key values we expect
    cells = {(c.value): c.coordinate for row in ws.iter_rows() for c in row if c.value is not None}
    # Period rendered
    assert any("2025-12-22 to 2025-12-22" in str(v) for v in cells)
    # Income total exists
    assert any(v == "Total income" for v in cells)
    # Net profit row present
    assert any(v == "NET PROFIT" for v in cells)
    # Estimated tax = 20% of net = 0.20 * (1500 - 14.95 - 2.60) = 296.49
    assert any(isinstance(v, (int, float)) and abs(v - 296.49) < 0.02 for v in cells)

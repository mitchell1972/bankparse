"""
Tests for the live tax-due forecast (services/tax_forecast.py) and
the accountant export ZIP (services/accountant_export.py).
"""
from __future__ import annotations

import base64
import datetime as _dt
import io
import json
import os
import secrets
import sys
import time
import zipfile

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

TEST_DB_PATH = "/tmp/test_bankparse_tax_forecast_export.db"


@pytest.fixture(autouse=True)
def _env(monkeypatch):
    monkeypatch.setenv(
        "HMRC_TOKEN_ENCRYPTION_KEY",
        base64.b64encode(secrets.token_bytes(32)).decode(),
    )
    monkeypatch.setenv("ENVIRONMENT", "development")


@pytest.fixture(autouse=True)
def clean_db():
    import database
    if os.path.exists(TEST_DB_PATH):
        os.unlink(TEST_DB_PATH)
    if database._sqlite_conn is not None:
        try: database._sqlite_conn.close()
        except Exception: pass
    database._sqlite_conn = None
    import sqlite3
    def _get_sqlite_test():
        if database._sqlite_conn is None:
            conn = sqlite3.connect(TEST_DB_PATH, timeout=10, check_same_thread=False)
            conn.execute("PRAGMA journal_mode=WAL")
            conn.execute("PRAGMA foreign_keys=ON")
            database._sqlite_conn = conn
        return database._sqlite_conn
    database._get_sqlite = _get_sqlite_test
    database.init_db()
    yield
    if database._sqlite_conn is not None:
        try: database._sqlite_conn.close()
        except Exception: pass
        database._sqlite_conn = None
    if os.path.exists(TEST_DB_PATH):
        os.unlink(TEST_DB_PATH)


def _seed_user(email: str = "u@example.com") -> int:
    import database as _db
    _db.create_user(email, "pwhash")
    return _db.get_user_by_email(email)["id"]


def _add_tx(user_id: int, **kwargs) -> int:
    import database as _db
    base = {"extracted_data_id": None,
            "date_iso": _dt.date.today().isoformat(),
            "description": "X", "amount": -10.0}
    base.update(kwargs)
    return _db.insert_ledger_transaction(user_id, **base)


# ---------------------------------------------------------------------------
# Tax forecast
# ---------------------------------------------------------------------------


def test_forecast_returns_zeros_for_empty_user():
    from services.tax_forecast import forecast_tax_due
    uid = _seed_user()
    f = forecast_tax_due(uid)
    assert f["combined"]["total_due"] == 0
    assert f["self_employment"]["profit"] == 0
    assert f["property"]["profit"] == 0


def test_forecast_below_personal_allowance_is_zero_tax():
    """Profit under £12,570 — no income tax due."""
    from services.tax_forecast import forecast_tax_due
    uid = _seed_user()
    _add_tx(uid, hmrc_category="se_turnover", amount=10000.0)
    _add_tx(uid, hmrc_category="se_office_expenses", amount=-1000.0)
    f = forecast_tax_due(uid)
    # Profit £9,000 < £12,570 PA → £0 tax
    assert f["combined"]["income_tax_due"] == 0
    assert f["combined"]["class_4_ni_due"] == 0
    assert any("personal allowance" in n.lower() for n in f["notes"])


def test_forecast_basic_rate_band_se():
    """Profit £30,000 - £12,570 PA = £17,430 in basic rate band = 20% = £3,486 tax.
    Plus Class 4 NI: (30000 - 12570) * 0.06 = £1,045.80."""
    from services.tax_forecast import forecast_tax_due
    uid = _seed_user()
    _add_tx(uid, hmrc_category="se_turnover", amount=30000.0)
    f = forecast_tax_due(uid)
    assert abs(f["combined"]["income_tax_due"] - 3486.0) < 1.0
    assert abs(f["combined"]["class_4_ni_due"] - 1045.80) < 1.0


def test_forecast_combined_se_plus_property():
    """SE profit £10k + property profit £20k = £30k total profit.
    Property profit doesn't attract Class 4 NI.
    Income tax = (30k - 12570) * 0.20 = £3,486.
    Class 4 NI = only on SE profit (£10k - £12,570 floor → £0, but
    NI is calculated on SE income alone, not combined). The forecast
    treats SE alone for NI."""
    from services.tax_forecast import forecast_tax_due
    uid = _seed_user()
    _add_tx(uid, hmrc_category="se_turnover", amount=10000.0)
    _add_tx(uid, hmrc_category="property_rent_income", amount=20000.0)
    f = forecast_tax_due(uid)
    # Combined profit £30k → £3,486 income tax across both streams
    assert abs(f["combined"]["income_tax_due"] - 3486.0) < 1.0
    # Class 4 NI = SE only. SE profit £10k < £12,570 floor → £0
    assert f["combined"]["class_4_ni_due"] == 0
    # Apportioned: SE share = 10k / 30k ≈ 33%, property = 67%
    assert f["self_employment"]["income_tax_due"] > 0
    assert f["property"]["income_tax_due"] > 0


def test_forecast_excludes_personal_transactions():
    from services.tax_forecast import forecast_tax_due
    import database as _db
    uid = _seed_user()
    _add_tx(uid, hmrc_category="se_turnover", amount=20000.0)
    personal = _add_tx(uid, hmrc_category="se_general_admin_costs", amount=-5000.0)
    _db.update_transaction_status(personal, receipt_status="excluded",
                                   exclusion_reason="personal")
    f = forecast_tax_due(uid)
    # Personal excluded → SE expenses = 0 → profit = 20k
    assert f["self_employment"]["expenses"] == 0
    assert f["self_employment"]["profit"] == 20000


def test_forecast_excludes_capital_items_from_revenue_expenses():
    """A £350 laptop marked is_capital=1 doesn't reduce SE profit (it's
    a capital allowance, not a revenue expense)."""
    from services.tax_forecast import forecast_tax_due
    import database as _db
    uid = _seed_user()
    _add_tx(uid, hmrc_category="se_turnover", amount=20000.0)
    cap = _add_tx(uid, hmrc_category="se_office_expenses", amount=-350.0)
    _db.update_transaction_status(cap, is_capital=1)
    f = forecast_tax_due(uid)
    assert f["self_employment"]["expenses"] == 0
    assert f["self_employment"]["profit"] == 20000


def test_forecast_counts_uncategorised_credits_as_provisional_income():
    """An uploaded statement whose transactions haven't been categorised
    yet should still produce a meaningful forecast — credits go into SE
    income provisionally, debits into SE expenses, with a caveat note.
    Pinned after Mitchell's 2026-05-20 audit."""
    from services.tax_forecast import forecast_tax_due
    uid = _seed_user()
    _add_tx(uid, hmrc_category=None, amount=20_000.0)
    _add_tx(uid, hmrc_category=None, amount=-3_000.0)
    f = forecast_tax_due(uid)
    assert f["provisional"]["uncategorised_income"] == 20_000.0
    assert f["provisional"]["uncategorised_expenses"] == 3_000.0
    # Flowed into SE buckets
    assert f["self_employment"]["income"] == 20_000.0
    assert f["self_employment"]["expenses"] == 3_000.0
    # Profit £17k > £12,570 → some income tax
    assert f["combined"]["income_tax_due"] > 0
    assert any("uncategorised" in n.lower() for n in f["notes"])


def test_forecast_uncategorised_handles_mitchell_real_data():
    """Real numbers from Mitchell's bank statement: credits £10,699.61,
    debits £8,439.72 → net £2,259.89, no tax due (under PA)."""
    from services.tax_forecast import forecast_tax_due
    uid = _seed_user()
    _add_tx(uid, hmrc_category=None, amount=10_699.61)
    _add_tx(uid, hmrc_category=None, amount=-8_439.72)
    f = forecast_tax_due(uid)
    assert abs(f["combined"]["profit"] - 2_259.89) < 0.01
    assert f["combined"]["income_tax_due"] == 0  # under PA
    assert any("uncategorised" in n.lower() for n in f["notes"])


def test_forecast_business_pct_proportional():
    """A 60% business expense reduces profit by 60% of the amount."""
    from services.tax_forecast import forecast_tax_due
    import database as _db
    uid = _seed_user()
    _add_tx(uid, hmrc_category="se_turnover", amount=20000.0)
    tx = _add_tx(uid, hmrc_category="se_motor_expenses", amount=-1000.0)
    _db.update_transaction_status(tx, business_pct=60)
    f = forecast_tax_due(uid)
    assert abs(f["self_employment"]["expenses"] - 600.0) < 0.01


# ---------------------------------------------------------------------------
# Capital prompt
# ---------------------------------------------------------------------------


def test_capital_prompt_fires_above_200_for_se_expense():
    from services.tax_forecast import should_prompt_capital
    assert should_prompt_capital({
        "amount": -350.0,
        "hmrc_category": "se_office_expenses",
    }) is True


def test_capital_prompt_silent_under_threshold():
    from services.tax_forecast import should_prompt_capital
    assert should_prompt_capital({
        "amount": -50.0,
        "hmrc_category": "se_office_expenses",
    }) is False


def test_capital_prompt_silent_for_income():
    from services.tax_forecast import should_prompt_capital
    assert should_prompt_capital({
        "amount": 500.0,  # positive = inflow
        "hmrc_category": "se_turnover",
    }) is False


def test_capital_prompt_silent_for_already_marked():
    from services.tax_forecast import should_prompt_capital
    assert should_prompt_capital({
        "amount": -350.0,
        "hmrc_category": "se_office_expenses",
        "is_capital": 1,
    }) is False


# ---------------------------------------------------------------------------
# Accountant export ZIP
# ---------------------------------------------------------------------------


def test_export_zip_contains_required_files():
    """ZIP layout (2026-05-20 redesign): the headline file is the Excel
    workbook + a README + grouped receipts. Raw CSVs live under data/."""
    from services.accountant_export import build_export_zip
    uid = _seed_user("export@example.com")
    _add_tx(uid, hmrc_category="adminCosts",
            description="AMAZON", amount=-42.99,
            date_iso="2026-08-04")
    zip_bytes = build_export_zip(uid, "export@example.com", period_label="Q2-2026")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    names = set(zf.namelist())
    # New top-level files
    assert "README.txt" in names
    assert "Accountant_Pack.xlsx" in names
    assert "summary.html" in names
    assert "manifest.json" in names
    # CSVs moved under data/
    assert "data/transactions.csv" in names
    assert "data/mismatch_report.csv" in names
    assert "data/reasoning_log.csv" in names


def test_export_zip_transactions_csv_includes_friendly_label_and_sa_box():
    """The CSV now carries human-friendly category + SA box mapping —
    the practice can import it into TaxCalc/CCH/IRIS without a translation
    table on the side."""
    from services.accountant_export import build_export_zip
    uid = _seed_user("cols@example.com")
    _add_tx(uid, hmrc_category="adminCosts",
            description="AMAZON", amount=-42.99,
            date_iso="2026-08-04")
    zip_bytes = build_export_zip(uid, "cols@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    with zf.open("data/transactions.csv") as fh:
        text = fh.read().decode()
    header = text.splitlines()[0]
    for col in ("hmrc_category", "hmrc_category_friendly",
                "sa_box_short", "sa_box_full",
                "vat_amount", "content_hash",
                "business_pct", "is_capital", "linked_receipt_ids"):
        assert col in header, f"missing column {col}"
    # Friendly label + SA box appear in row data
    assert "AMAZON" in text
    assert "SA103S Box 18" in text  # adminCosts → SA103S Box 18
    assert "office costs" in text.lower()  # friendly label


def test_export_zip_manifest_lists_all_file_hashes():
    from services.accountant_export import build_export_zip
    uid = _seed_user("manifest@example.com")
    _add_tx(uid, hmrc_category="adminCosts", amount=-50.0)
    zip_bytes = build_export_zip(uid, "manifest@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    manifest = json.loads(zf.read("manifest.json"))
    file_hashes = manifest["file_hashes"]
    for f in ("README.txt", "Accountant_Pack.xlsx", "summary.html",
              "data/transactions.csv", "data/mismatch_report.csv",
              "data/reasoning_log.csv"):
        assert f in file_hashes, f"missing hash for {f}"
        assert len(file_hashes[f]) == 64  # SHA-256


def test_export_zip_mismatch_report_only_includes_unmatched():
    """A transaction with a linked receipt should NOT appear in the
    mismatch report."""
    from services.accountant_export import build_export_zip
    import database as _db
    uid = _seed_user("mm@example.com")
    tx_matched = _add_tx(uid, hmrc_category="adminCosts",
                         description="MATCHED", amount=-42.99)
    tx_missing = _add_tx(uid, hmrc_category="travelCosts",
                         description="MISSING", amount=-25.00)
    rc = _db.insert_ledger_receipt(
        uid, extracted_data_id=None, file_path=None, source_filename="r.pdf",
        store_name="A", date_iso=_dt.date.today().isoformat(),
        total_amount=42.99, tax_amount=7.16,
    )
    _db.insert_ledger_link(transaction_id=tx_matched, receipt_id=rc,
                           match_strategy="exact", confidence=100)

    zip_bytes = build_export_zip(uid, "mm@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    mismatch = zf.read("data/mismatch_report.csv").decode()
    assert "MISSING" in mismatch
    assert "MATCHED" not in mismatch


def test_export_zip_readme_names_the_client_and_period():
    """README.txt is the first file the accountant opens. It must name
    the client + period clearly so they know which file they're holding."""
    from services.accountant_export import build_export_zip
    uid = _seed_user("readme@example.com")
    _add_tx(uid, hmrc_category="adminCosts", amount=-10.0)
    zip_bytes = build_export_zip(
        uid, "readme@example.com",
        period_label="Q2 2026-27 (Jul-Sep)",
        client_name="Mitoba Property Services Ltd",
    )
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    readme = zf.read("README.txt").decode()
    assert "Mitoba Property Services Ltd" in readme
    assert "Q2 2026-27 (Jul-Sep)" in readme
    assert "Accountant_Pack.xlsx" in readme
    assert "SA103" in readme  # tells the accountant the box mapping exists


def test_export_xlsx_has_seven_sheets_with_box_mapping():
    """The workbook is what an accountant actually opens. It must have all
    seven sheets, and the Tax Return Boxes sheet must contain SA box refs."""
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("xlsx@example.com")
    _add_tx(uid, hmrc_category="adminCosts",
            description="STAPLES", amount=-15.99)
    _add_tx(uid, hmrc_category="turnover",
            description="CLIENT INVOICE", amount=2000.0)
    zip_bytes = build_export_zip(uid, "xlsx@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    xlsx_bytes = zf.read("Accountant_Pack.xlsx")
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    expected = {
        "Cover", "Tax Return Boxes", "Transactions", "Missing Receipts",
        "Receipt Inventory", "VAT Register", "Reasoning Log",
    }
    assert expected.issubset(set(wb.sheetnames)), (
        f"missing sheets: {expected - set(wb.sheetnames)}"
    )
    # Box mapping sheet has SA103S box numbers
    boxes_text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Tax Return Boxes"].iter_rows()
    )
    assert "SA103S Box 9" in boxes_text   # turnover
    assert "SA103S Box 18" in boxes_text  # adminCosts
    assert "Phone, stationery" in boxes_text  # friendly label


def test_export_xlsx_cover_includes_client_name():
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("cover@example.com")
    _add_tx(uid, hmrc_category="adminCosts", amount=-50.0)
    zip_bytes = build_export_zip(
        uid, "cover@example.com",
        client_name="Acme Plumbing Ltd",
    )
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    cover_text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Cover"].iter_rows()
    )
    assert "Acme Plumbing Ltd" in cover_text
    assert "Tax Return Boxes" in cover_text  # instructions point there


def test_export_zip_receipts_grouped_by_hmrc_category():
    """Receipts attached to a category-X transaction should land in
    receipts/X/ inside the ZIP. Orphan receipts go to receipts/_orphan/."""
    from services.accountant_export import build_export_zip
    import database as _db, tempfile
    uid = _seed_user("grouped@example.com")
    # Create a real on-disk receipt file so the export actually attaches it
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(b"%PDF-1.4 fake")
        tmp_path = tmp.name
    try:
        tx = _add_tx(uid, hmrc_category="adminCosts",
                     description="STAPLES", amount=-15.99)
        rc_matched = _db.insert_ledger_receipt(
            uid, extracted_data_id=None, file_path=tmp_path,
            source_filename="staples.pdf",
            store_name="Staples", date_iso="2026-08-04",
            total_amount=15.99, tax_amount=2.66,
        )
        _db.insert_ledger_link(
            transaction_id=tx, receipt_id=rc_matched,
            match_strategy="exact", confidence=100,
        )
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp2:
            tmp2.write(b"%PDF-1.4 orphan")
            orphan_path = tmp2.name
        try:
            _db.insert_ledger_receipt(
                uid, extracted_data_id=None, file_path=orphan_path,
                source_filename="orphan.pdf",
                store_name="Mystery", date_iso="2026-08-05",
                total_amount=9.99, tax_amount=1.66,
            )
            zip_bytes = build_export_zip(uid, "grouped@example.com")
        finally:
            import os
            os.unlink(orphan_path)
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
        names = zf.namelist()
        # Matched receipt is under receipts/adminCosts/
        assert any(n.startswith("receipts/adminCosts/") for n in names), \
            f"no receipts/adminCosts/ in {[n for n in names if n.startswith('receipts/')]}"
        # Orphan goes to receipts/_orphan/
        assert any(n.startswith("receipts/_orphan/") for n in names), \
            f"no receipts/_orphan/ in {[n for n in names if n.startswith('receipts/')]}"
    finally:
        import os
        os.unlink(tmp_path)


def test_export_zip_endpoint_returns_application_zip():
    from app import app
    import database as _db
    client = TestClient(app, raise_server_exceptions=False)
    client.__enter__()
    client.get("/login")
    csrf = client.cookies.get("bp_csrf", "")
    client.post("/api/register",
                json={"email": "zipuser@example.com", "password": "password12345"},
                headers={"X-CSRF-Token": csrf})
    user = _db.get_user_by_email("zipuser@example.com")
    _db.mark_email_verified(user["id"])
    _db.update_user(user["id"], subscription_status="trialing",
                    stripe_subscription_id="sub_zip",
                    trial_end_at=time.time() + 7*86400)

    r = client.get("/api/accountant-export?period=Q2-2026")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/zip"
    assert "Q2-2026" in r.headers.get("content-disposition", "")


def test_export_endpoint_requires_auth():
    from app import app
    client = TestClient(app, raise_server_exceptions=False)
    client.__enter__()
    r = client.get("/api/accountant-export")
    assert r.status_code == 401


def test_tax_forecast_endpoint_returns_shape():
    from app import app
    import database as _db
    client = TestClient(app, raise_server_exceptions=False)
    client.__enter__()
    client.get("/login")
    csrf = client.cookies.get("bp_csrf", "")
    client.post("/api/register",
                json={"email": "forecast@example.com", "password": "password12345"},
                headers={"X-CSRF-Token": csrf})
    user = _db.get_user_by_email("forecast@example.com")
    _db.mark_email_verified(user["id"])
    _db.update_user(user["id"], subscription_status="trialing",
                    stripe_subscription_id="sub_fc",
                    trial_end_at=time.time() + 7*86400)

    r = client.get("/api/tax-forecast")
    assert r.status_code == 200
    body = r.json()
    assert "self_employment" in body
    assert "property" in body
    assert "combined" in body
    assert "tax_year" in body


# ---------------------------------------------------------------------------
# Accountant pack — Action Items sheet (priority sheet — first stop for the
# practice firm after the Cover)
# ---------------------------------------------------------------------------


def test_action_items_sheet_shows_uncategorised_as_high_severity():
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("ai-uncat@example.com")
    _add_tx(uid, hmrc_category=None, description="MYSTERY", amount=-50.0)
    zip_bytes = build_export_zip(uid, "ai-uncat@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    assert "Action Items" in wb.sheetnames
    text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Action Items"].iter_rows()
    )
    assert "MYSTERY" in text
    assert "HIGH" in text
    assert "No HMRC category" in text


def test_action_items_sheet_flags_low_confidence_categorisation():
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("ai-lowconf@example.com")
    _add_tx(uid, hmrc_category="adminCosts",
            hmrc_category_confidence=35,
            description="STAPLES MAYBE", amount=-15.99)
    zip_bytes = build_export_zip(uid, "ai-lowconf@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Action Items"].iter_rows()
    )
    assert "STAPLES MAYBE" in text
    assert "Low confidence" in text


def test_action_items_sheet_flags_large_expense_with_no_receipt():
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("ai-noreceipt@example.com")
    _add_tx(uid, hmrc_category="adminCosts",
            hmrc_category_confidence=90,
            description="BIG SPEND", amount=-250.00)
    zip_bytes = build_export_zip(uid, "ai-noreceipt@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Action Items"].iter_rows()
    )
    assert "BIG SPEND" in text
    assert "no receipt" in text.lower()


def test_action_items_sheet_says_all_clear_when_no_issues():
    """The accountant should see a clear ✓ if everything's tidy — not a
    misleading blank sheet."""
    from services.accountant_export import build_export_zip
    import database as _db
    from openpyxl import load_workbook
    uid = _seed_user("ai-clean@example.com")
    # Categorised + high confidence + receipt-backed + small amount = no issues
    tx = _add_tx(uid, hmrc_category="adminCosts",
                 hmrc_category_confidence=95,
                 description="CLEAN", amount=-9.99)
    rc = _db.insert_ledger_receipt(
        uid, extracted_data_id=None, file_path=None, source_filename="r.pdf",
        store_name="Clean Co", date_iso=_dt.date.today().isoformat(),
        total_amount=9.99, tax_amount=1.66,
    )
    _db.insert_ledger_link(transaction_id=tx, receipt_id=rc,
                           match_strategy="exact", confidence=100)
    zip_bytes = build_export_zip(uid, "ai-clean@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Action Items"].iter_rows()
    )
    assert "No issues detected" in text


# ---------------------------------------------------------------------------
# Trial Balance sheet — practice-tool import format
# ---------------------------------------------------------------------------


def test_trial_balance_uncategorised_expense_shows_friendly_label():
    """Sentinel "uncategorised" bucket must render as a clear
    "Uncategorised expenses — review needed" not "Unrecognised code:
    uncategorised". Mitchell-quote: 'this feels DIY'."""
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("uncat-label@example.com")
    _add_tx(uid, hmrc_category=None, description="MYSTERY", amount=-89.50)
    zip_bytes = build_export_zip(uid, "uncat-label@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Trial Balance"].iter_rows()
    )
    assert "Uncategorised expenses — review needed" in text
    assert "Unrecognised code" not in text


def test_trial_balance_sheet_has_correct_debit_credit_columns():
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("tb@example.com")
    _add_tx(uid, hmrc_category="turnover",
            description="CLIENT", amount=1000.00)
    _add_tx(uid, hmrc_category="adminCosts",
            description="OFFICE", amount=-150.00)
    zip_bytes = build_export_zip(uid, "tb@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    assert "Trial Balance" in wb.sheetnames
    ws = wb["Trial Balance"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    assert headers == (
        "Nominal code", "Description", "Debit (GBP)", "Credit (GBP)", "Net (GBP)",
    )
    # Build a lookup by nominal code
    by_code = {r[0]: r for r in rows[1:] if r[0] and r[0] != "TOTALS"}
    # turnover should be on the credit side
    assert by_code["turnover"][3] == 1000.00
    assert by_code["turnover"][2] in (0, None)
    # adminCosts should be on the debit side
    assert by_code["adminCosts"][2] == 150.00
    assert by_code["adminCosts"][3] in (0, None)


def test_trial_balance_totals_reconcile_to_cover():
    """The TB totals row MUST match the Cover's Income/Expenses/Net so the
    accountant's first sanity check passes. Tested with both income and
    expense categories present."""
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("tb-recon@example.com")
    _add_tx(uid, hmrc_category="turnover",
            description="A", amount=2000.00)
    _add_tx(uid, hmrc_category="turnover",
            description="B", amount=1500.00)
    _add_tx(uid, hmrc_category="adminCosts",
            description="C", amount=-300.00)
    zip_bytes = build_export_zip(uid, "tb-recon@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    tb_rows = list(wb["Trial Balance"].iter_rows(values_only=True))
    totals_row = next(r for r in tb_rows if r[1] == "TOTALS")
    assert totals_row[2] == 300.00   # debit total = expenses
    assert totals_row[3] == 3500.00  # credit total = income
    assert totals_row[4] == 3200.00  # net = income - expenses


# ---------------------------------------------------------------------------
# Period filter — actually filters rows, not just labels the cover
# ---------------------------------------------------------------------------


def test_period_filter_drops_transactions_outside_q2_2026_27():
    """Period dropdown set to Q2 2026-27 must exclude April + October rows."""
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("period@example.com")
    _add_tx(uid, hmrc_category="adminCosts", date_iso="2026-04-15",
            description="APR_PRE_Q2", amount=-10.0)
    _add_tx(uid, hmrc_category="adminCosts", date_iso="2026-08-15",
            description="AUG_IN_Q2", amount=-20.0)
    _add_tx(uid, hmrc_category="adminCosts", date_iso="2026-10-20",
            description="OCT_AFTER_Q2", amount=-30.0)
    zip_bytes = build_export_zip(
        uid, "period@example.com",
        period_label="Q2 2026-27 (Jul-Sep)",
    )
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    # Transactions CSV should only contain the in-period row
    tx_csv = zf.read("data/transactions.csv").decode()
    assert "AUG_IN_Q2" in tx_csv
    assert "APR_PRE_Q2" not in tx_csv
    assert "OCT_AFTER_Q2" not in tx_csv
    # Cover marks the pack as filtered
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    cover_text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Cover"].iter_rows()
    )
    assert "FILTERED" in cover_text
    assert "Filtered to period" in cover_text


def test_period_filter_none_keeps_all_transactions():
    """Empty period label → no filter, pack shows all-time."""
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("nofilter@example.com")
    _add_tx(uid, hmrc_category="adminCosts", date_iso="2020-01-01",
            description="OLD", amount=-5.0)
    _add_tx(uid, hmrc_category="adminCosts", date_iso="2027-01-01",
            description="FUTURE", amount=-7.0)
    zip_bytes = build_export_zip(uid, "nofilter@example.com")
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    tx_csv = zf.read("data/transactions.csv").decode()
    assert "OLD" in tx_csv
    assert "FUTURE" in tx_csv
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    cover_text = "\n".join(
        " ".join(str(c.value or "") for c in row)
        for row in wb["Cover"].iter_rows()
    )
    assert "ALL TIME" in cover_text


def test_period_filter_cover_totals_match_filtered_transactions():
    """The Cover's headline totals must reconcile to the filtered tx set —
    otherwise the accountant sees mismatched numbers and loses trust."""
    from services.accountant_export import build_export_zip
    from openpyxl import load_workbook
    uid = _seed_user("recon@example.com")
    # In period
    _add_tx(uid, hmrc_category="turnover", date_iso="2026-08-15",
            description="IN_INCOME", amount=1000.00)
    _add_tx(uid, hmrc_category="adminCosts", date_iso="2026-08-16",
            description="IN_EXPENSE", amount=-200.00)
    # Out of period (must NOT count)
    _add_tx(uid, hmrc_category="turnover", date_iso="2026-04-01",
            description="OUT_INCOME", amount=9999.00)
    zip_bytes = build_export_zip(
        uid, "recon@example.com",
        period_label="Q2 2026-27 (Jul-Sep)",
    )
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    wb = load_workbook(io.BytesIO(zf.read("Accountant_Pack.xlsx")))
    # Walk the Cover's right-hand totals block. Cell values are raw floats —
    # the £ symbol is a number_format applied at render time.
    cover = wb["Cover"]
    totals_by_label: dict[str, float] = {}
    for row in cover.iter_rows():
        if len(row) < 4:
            continue
        label = row[2].value
        val = row[3].value
        if isinstance(label, str) and isinstance(val, (int, float)):
            totals_by_label[label.strip()] = float(val)

    assert totals_by_label.get("Income (gross)") == 1000.00, totals_by_label
    assert totals_by_label.get("Expenses (gross)") == 200.00, totals_by_label
    assert totals_by_label.get("Net profit") == 800.00, totals_by_label
    # Out-of-period income must NOT inflate the totals
    assert 9999.00 not in totals_by_label.values()

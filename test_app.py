"""
Test suite for BankParse — tests parsers, XLSX export, database, OTP, and API endpoints.
"""
import os
import csv
import time
import pytest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient

# Set test database path before importing app
os.environ["BANKPARSE_DB_PATH"] = "/tmp/bankparse_test.db"

from app import app
from database import (
    get_usage, save_usage, increment_usage,
    store_otp, verify_otp, cleanup_expired_otps,
    track_output_file, get_stale_output_files, remove_output_file_record,
)
from otp import generate_otp
from parsers.csv_parser import parse_csv
from parsers.xlsx_exporter import export_to_xlsx, export_receipt_to_xlsx
from parsers.receipt_parser import parse_receipt_text


@pytest.fixture(autouse=True)
def clean_db():
    """Reset the test database before each test."""
    from database import _execute
    _execute("DELETE FROM sessions")
    _execute("DELETE FROM otp_codes")
    _execute("DELETE FROM output_files")
    yield


@pytest.fixture
def client():
    return TestClient(app)


def create_sample_csv(path: str):
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Date", "Description", "Money Out", "Money In", "Balance"])
        writer.writerow(["15/01/2025", "TESCO STORES 3217", "45.67", "", "1954.33"])
        writer.writerow(["15/01/2025", "SALARY - ACME LTD", "", "2850.00", "4804.33"])
        writer.writerow(["16/01/2025", "DIRECT DEBIT - VODAFONE", "32.00", "", "4772.33"])
        writer.writerow(["17/01/2025", "TRANSFER - J SMITH", "", "150.00", "4922.33"])
        writer.writerow(["18/01/2025", "AMAZON UK MARKETPLACE", "89.99", "", "4832.34"])
        writer.writerow(["19/01/2025", "STANDING ORDER - RENT", "750.00", "", "4082.34"])
        writer.writerow(["20/01/2025", "CONTACTLESS - COSTA COFFEE", "4.50", "", "4077.84"])
        writer.writerow(["21/01/2025", "FASTER PAYMENT - FREELANCE", "", "500.00", "4577.84"])
        writer.writerow(["22/01/2025", "DIRECT DEBIT - COUNCIL TAX", "156.00", "", "4421.84"])
        writer.writerow(["23/01/2025", "CARD PAYMENT - SAINSBURYS", "62.30", "", "4359.54"])
        writer.writerow(["24/01/2025", "INTEREST PAID", "", "1.23", "4360.77"])
        writer.writerow(["25/01/2025", "ATM WITHDRAWAL - HSBC", "200.00", "", "4160.77"])


def create_sample_csv_lloyds(path: str):
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Transaction Date", "Transaction Type", "Transaction Description", "Debit Amount", "Credit Amount", "Balance"])
        writer.writerow(["01/02/2025", "DD", "NETFLIX.COM", "15.99", "", "3200.01"])
        writer.writerow(["02/02/2025", "FPO", "HMRC TAX REFUND", "", "350.00", "3550.01"])
        writer.writerow(["03/02/2025", "CPT", "PRET A MANGER", "7.45", "", "3542.56"])
        writer.writerow(["04/02/2025", "BGC", "SALARY", "", "3100.00", "6642.56"])
        writer.writerow(["05/02/2025", "DD", "SCOTTISH POWER", "89.00", "", "6553.56"])


# --- Parser Tests ---

def test_csv_parser(tmp_path):
    path = str(tmp_path / "barclays.csv")
    create_sample_csv(path)
    result = parse_csv(path)

    assert len(result["transactions"]) == 12
    assert result["summary"]["total_transactions"] == 12

    salary = [t for t in result["transactions"] if "ACME" in t["description"]]
    assert len(salary) == 1
    assert salary[0]["amount"] == 2850.00


def test_csv_parser_lloyds(tmp_path):
    path = str(tmp_path / "lloyds.csv")
    create_sample_csv_lloyds(path)
    result = parse_csv(path)

    assert len(result["transactions"]) == 5
    assert result["summary"]["total_transactions"] == 5


def test_xlsx_export(tmp_path):
    csv_path = str(tmp_path / "barclays.csv")
    create_sample_csv(csv_path)
    data = parse_csv(csv_path)

    xlsx_path = str(tmp_path / "output.xlsx")
    export_to_xlsx(data, xlsx_path)

    assert os.path.getsize(xlsx_path) > 0

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    assert "Transactions" in wb.sheetnames
    assert "Summary" in wb.sheetnames


def test_full_pipeline(tmp_path):
    csv_path = str(tmp_path / "statement.csv")
    create_sample_csv(csv_path)
    result = parse_csv(csv_path)

    xlsx_path = str(tmp_path / "output.xlsx")
    export_to_xlsx(result, xlsx_path)

    assert os.path.exists(xlsx_path)
    assert os.path.getsize(xlsx_path) > 0
    assert len(result["transactions"]) == 12


def test_receipt_parser():
    sample_receipt = """
TESCO STORES LTD
Express Manchester
14 Deansgate M3 2BA
Tel: 0345 6779012

15/03/2025  14:32

Whole Milk 2L          1.55
Hovis Bread 800g       1.35
Chicken Breast 500g    3.80
Basmati Rice 1kg       1.89
2x Heinz Beans 415g    2.00
Lurpak Butter 250g     2.25
6x Eggs Free Range     1.95
Red Onions 3pk         0.79
Fairy Liquid 500ml     1.65
PG Tips 80s            3.15

SUBTOTAL              20.38
VAT                    0.00
TOTAL                 20.38

CARD PAYMENT          20.38
Visa Debit ****7742

Clubcard Points: 20
THANK YOU FOR SHOPPING AT TESCO
"""
    result = parse_receipt_text(sample_receipt)

    assert len(result["items"]) == 10
    assert result["metadata"]["store_name"] == "TESCO STORES LTD"
    assert result["metadata"]["date"] == "2025-03-15"
    assert result["totals"].get("total") == 20.38


def test_receipt_parser_sainsburys():
    sample_receipt = """
Sainsbury's
Local Wandsworth
22/03/2025

Bananas Loose          0.65
Semi Skimmed Milk 1L   1.10
Warburtons Toastie     1.30
Cathedral Cheddar      3.00
Chicken Thighs 1kg     4.50
Broccoli               0.75
Cherry Tomatoes 250g   1.25
Muller Corner x4       2.50
Persil Liquid 1.5L     5.00

Subtotal              20.05
Savings               -1.50
TOTAL                 18.55

Paid by Card          18.55
Nectar Points: 18
"""
    result = parse_receipt_text(sample_receipt)

    assert len(result["items"]) == 9
    assert result["totals"].get("total") == 18.55
    assert result["metadata"]["date"] == "2025-03-22"


def test_receipt_xlsx_export(tmp_path):
    data = {
        "items": [
            {"description": "Milk", "quantity": 1, "unit_price": 1.55, "total_price": 1.55},
            {"description": "Bread", "quantity": 2, "unit_price": 1.20, "total_price": 2.40},
            {"description": "Eggs", "quantity": 1, "unit_price": 1.95, "total_price": 1.95},
        ],
        "totals": {"subtotal": 5.90, "total": 5.90},
        "metadata": {"store_name": "Test Store", "date": "2025-03-15", "item_count": 3, "currency": "GBP"},
    }

    xlsx_path = str(tmp_path / "receipt.xlsx")
    export_receipt_to_xlsx(data, xlsx_path)

    assert os.path.getsize(xlsx_path) > 0

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path)
    assert "Receipt Items" in wb.sheetnames
    ws = wb["Receipt Items"]
    assert "Test Store" in str(ws["A1"].value)


# --- Database Tests ---

def test_database_usage_crud():
    usage = get_usage("test_session_1")
    assert usage["statements"] == 0
    assert usage["receipts"] == 0

    usage["statements"] = 3
    usage["email"] = "test@example.com"
    save_usage("test_session_1", usage)

    loaded = get_usage("test_session_1")
    assert loaded["statements"] == 3
    assert loaded["email"] == "test@example.com"


def test_database_increment_usage():
    increment_usage("test_session_2", "statement")
    increment_usage("test_session_2", "statement")
    increment_usage("test_session_2", "receipt")

    usage = get_usage("test_session_2")
    assert usage["statements"] == 2
    assert usage["receipts"] == 1


def test_database_empty_session():
    usage = get_usage("")
    assert usage["statements"] == 0
    assert usage["stripe_customer_id"] is None


# --- OTP Tests ---

def test_otp_generation():
    code = generate_otp()
    assert len(code) == 6
    assert code.isdigit()


def test_otp_store_and_verify():
    store_otp("user@test.com", "123456", "session_abc")
    result = verify_otp("user@test.com", "123456")
    assert result == "session_abc"


def test_otp_single_use():
    store_otp("user2@test.com", "654321", "session_xyz")
    assert verify_otp("user2@test.com", "654321") == "session_xyz"
    assert verify_otp("user2@test.com", "654321") is None


def test_otp_wrong_code():
    store_otp("user3@test.com", "111111", "session_123")
    assert verify_otp("user3@test.com", "999999") is None


def test_otp_invalidates_previous():
    store_otp("user4@test.com", "111111", "session_a")
    store_otp("user4@test.com", "222222", "session_b")
    assert verify_otp("user4@test.com", "111111") is None
    assert verify_otp("user4@test.com", "222222") == "session_b"


# --- Output File Tracking Tests ---

def test_output_file_tracking():
    track_output_file("test_file.xlsx")
    stale = get_stale_output_files(max_age_seconds=0)
    assert "test_file.xlsx" in stale

    remove_output_file_record("test_file.xlsx")
    stale = get_stale_output_files(max_age_seconds=0)
    assert "test_file.xlsx" not in stale


# --- API Tests ---

def test_health_endpoint(client):
    response = client.get("/api/health")
    assert response.status_code == 200
    assert response.json()["status"] == "ok"
    assert "version" in response.json()


def test_home_page(client):
    response = client.get("/")
    assert response.status_code == 200


def test_parse_invalid_file(client):
    response = client.post("/api/parse", files={"file": ("test.exe", b"data", "application/octet-stream")})
    assert response.status_code in (400, 403)  # 403 if auth required, 400 if bad file type


def test_usage_endpoint(client):
    response = client.get("/api/usage")
    assert response.status_code == 200
    data = response.json()
    assert "statements_used" in data
    assert "receipts_used" in data
    assert "has_subscription" in data


def test_config_endpoint(client):
    response = client.get("/api/config")
    assert response.status_code == 200
    data = response.json()
    assert "plans" in data
    assert "stripe_publishable_key" in data


# ── Admin users endpoint tests ────────────────────────────────────────

def test_admin_users_unauthenticated(client):
    """Should reject unauthenticated requests."""
    response = client.get("/api/admin/users")
    assert response.status_code == 401


def test_admin_users_non_admin(client):
    """Should reject non-admin authenticated users."""
    with patch("app.get_current_user", return_value={"id": 99, "email": "regular@example.com"}):
        response = client.get("/api/admin/users")
        assert response.status_code == 403


def test_admin_users_returns_all_users(client):
    """Admin should see all registered users."""
    mock_users = [
        {"id": 1, "email": "a@example.com", "subscription_status": "active", "stripe_customer_id": "cus_1", "created_at": 1700000000},
        {"id": 2, "email": "b@example.com", "subscription_status": None, "stripe_customer_id": None, "created_at": 1700000001},
        {"id": 3, "email": "c@example.com", "subscription_status": None, "stripe_customer_id": None, "created_at": 1700000002},
    ]

    with patch("app.get_current_user", return_value={"id": 1, "email": "mitchell_agoma@yahoo.co.uk"}), \
         patch("app._fetchall_dicts", return_value=mock_users):
        response = client.get("/api/admin/users")
        assert response.status_code == 200
        data = response.json()
        assert data["total_users"] == 3
        assert len(data["users"]) == 3
        user = data["users"][0]
        assert "email" in user
        assert "subscription_status" in user
        assert "created_at" in user


# ── Admin page route tests ────────────────────────────────────────────

def test_admin_page_unauthenticated(client):
    """Should redirect to login when not authenticated."""
    response = client.get("/admin", follow_redirects=False)
    assert response.status_code == 302
    assert "/login" in response.headers["location"]


def test_admin_page_non_admin(client):
    """Should redirect to home for non-admin users."""
    with patch("app.get_current_user", return_value={"id": 99, "email": "regular@example.com"}):
        response = client.get("/admin", follow_redirects=False)
        assert response.status_code == 302
        assert response.headers["location"] == "/"


def test_admin_page_renders_for_admin(client):
    """Should render the admin dashboard for admin users."""
    with patch("app.get_current_user", return_value={"id": 1, "email": "mitchell_agoma@yahoo.co.uk"}):
        response = client.get("/admin")
        assert response.status_code == 200
        assert "Admin Dashboard" in response.text


# ── Comprehensive admin access control tests ──────────────────────────

def test_admin_page_blocked_for_similar_emails(client):
    """Emails that look similar but are NOT the admin should be blocked."""
    blocked_emails = [
        "mitchell_agoma@yahoo.com",        # wrong TLD
        "mitchell_agoma@gmail.com",        # different provider
        "mitchell_agoma@yahoo.co.uk ",     # trailing space
        " mitchell_agoma@yahoo.co.uk",     # leading space
        "admin@bankscanai.com",            # site admin guess
        "test@bankparse.com",              # test account
        "mitchell@yahoo.co.uk",            # partial name match
        "",                                # empty email
    ]
    for email in blocked_emails:
        with patch("app.get_current_user", return_value={"id": 99, "email": email}):
            response = client.get("/admin", follow_redirects=False)
            assert response.status_code == 302, f"Expected redirect for '{email}', got {response.status_code}"
            assert response.headers["location"] == "/", f"Expected redirect to / for '{email}'"


def test_admin_api_blocked_for_similar_emails(client):
    """The /api/admin/users endpoint should also block non-admin emails."""
    blocked_emails = [
        "mitchell_agoma@yahoo.com",
        "admin@bankscanai.com",
        "someone@example.com",
        "",
    ]
    for email in blocked_emails:
        with patch("app.get_current_user", return_value={"id": 99, "email": email}):
            response = client.get("/api/admin/users")
            assert response.status_code == 403, f"Expected 403 for '{email}', got {response.status_code}"


def test_admin_api_subscribers_blocked_for_non_admin(client):
    """The /api/admin/subscribers endpoint should also block non-admin emails."""
    with patch("app.get_current_user", return_value={"id": 99, "email": "hacker@evil.com"}):
        response = client.get("/api/admin/subscribers")
        assert response.status_code == 403


def test_admin_page_allowed_case_insensitive(client):
    """Admin email check should be case-insensitive."""
    with patch("app.get_current_user", return_value={"id": 1, "email": "Mitchell_Agoma@Yahoo.CO.UK"}):
        response = client.get("/admin")
        assert response.status_code == 200
        assert "Admin Dashboard" in response.text


def test_admin_api_allowed_case_insensitive(client):
    """Admin API check should be case-insensitive."""
    mock_users = [{"id": 1, "email": "a@b.com", "subscription_status": None, "stripe_customer_id": None, "created_at": 1700000000}]
    with patch("app.get_current_user", return_value={"id": 1, "email": "MITCHELL_AGOMA@YAHOO.CO.UK"}), \
         patch("app._fetchall_dicts", return_value=mock_users):
        response = client.get("/api/admin/users")
        assert response.status_code == 200


# ── Admin delete user tests ───────────────────────────────────────────

def _csrf_delete(client, url, **kwargs):
    """Helper: GET a page to obtain a CSRF cookie, then DELETE with that token."""
    get_res = client.get("/api/health")
    csrf_token = get_res.cookies.get("bp_csrf", "")
    return client.delete(
        url,
        cookies={"bp_csrf": csrf_token},
        headers={"X-CSRF-Token": csrf_token},
        **kwargs,
    )


def test_admin_delete_unauthenticated(client):
    """Should reject unauthenticated delete requests."""
    response = _csrf_delete(client, "/api/admin/users/99")
    assert response.status_code == 401


def test_admin_delete_non_admin(client):
    """Non-admin users cannot delete."""
    with patch("app.get_current_user", return_value={"id": 5, "email": "regular@example.com"}):
        response = _csrf_delete(client, "/api/admin/users/99")
        assert response.status_code == 403


def test_admin_delete_cannot_delete_self(client):
    """Admin cannot delete their own account."""
    with patch("app.get_current_user", return_value={"id": 1, "email": "mitchell_agoma@yahoo.co.uk"}):
        response = _csrf_delete(client, "/api/admin/users/1")
        assert response.status_code == 400
        assert "own account" in response.json()["detail"]


def test_admin_delete_user_success(client):
    """Admin can delete another user."""
    with patch("app.get_current_user", return_value={"id": 1, "email": "mitchell_agoma@yahoo.co.uk"}), \
         patch("app.delete_user") as mock_delete:
        response = _csrf_delete(client, "/api/admin/users/99")
        assert response.status_code == 200
        assert response.json()["ok"] is True
        mock_delete.assert_called_once_with(99)


# ── Admin button visibility test ──────────────────────────────────────

def test_admin_link_in_main_page(client):
    """The main page should contain a hidden admin link element."""
    with patch("app.get_current_user", return_value={"id": 1, "email": "mitchell_agoma@yahoo.co.uk"}):
        response = client.get("/")
        assert response.status_code == 200
        assert 'id="headerAdminLink"' in response.text
        assert 'href="/admin"' in response.text

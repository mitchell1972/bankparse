"""
Multi-sheet XLSX workbook for the Accountant Pack.

Before this module existed the accountant export was a pile of raw CSVs
the practice had to wrangle. Now it's a single workbook the accountant
can open in Excel, scan the cover sheet, and transcribe box totals
directly onto SA103S/SA105.

Sheets:

  1. Cover                — Practice info, period, totals, where to start
  2. Tax Return Boxes     — Each category mapped to SA103S / SA105 box +
                            ready-to-transcribe total. THE sheet that
                            saves the accountant 80% of the work.
  3. Transactions         — Full ledger with HMRC-friendly columns
  4. Missing Receipts     — Expenses with no receipt + reason
  5. Receipt Inventory    — Every receipt (matched / orphan)
  6. VAT Register         — Lines with VAT recorded
  7. Reasoning Log        — AI confidence + reasoning per category decision

Returns ``bytes`` so the caller can drop it straight into the ZIP.
"""
from __future__ import annotations

import datetime as _dt
import io
import json

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from hmrc.schemas import categories as _cats


# ---------------------------------------------------------------------------
# Curated metadata — canonical category → (human label, SA103/SA105 box,
# brief HMRC manual ref). Sourced from the 2026-27 SA103S, SA103F, and SA105
# published notes.
#
# IMPORTANT: keys MUST come from hmrc.schemas.categories — the architecture
# test (tests/hmrc/test_architecture.py) bans hardcoded category literals
# outside that canonical module.
# ---------------------------------------------------------------------------

# (label, sa_box_short, sa_box_full, hmrc_manual_ref, is_income)
_SE_META: dict[str, tuple[str, str, str, str, bool]] = {
    _cats.SE_INCOME: (
        "Turnover (sales)",
        "SA103S Box 9", "SA103F Box 15",
        "BIM40000", True,
    ),
    _cats.SE_OTHER_INCOME: (
        "Any other business income",
        "SA103S Box 10", "SA103F Box 16",
        "BIM40000", True,
    ),
    _cats.SE_EXPENSE_COST_OF_GOODS: (
        "Cost of goods bought for resale",
        "SA103S Box 11", "SA103F Box 17",
        "BIM38200", False,
    ),
    _cats.SE_EXPENSE_CIS: (
        "CIS payments to subcontractors",
        "SA103S Box 11", "SA103F Box 18",
        "BIM43500", False,
    ),
    _cats.SE_EXPENSE_STAFF: (
        "Wages, salaries and other staff costs",
        "SA103S Box 13", "SA103F Box 19",
        "BIM47000", False,
    ),
    _cats.SE_EXPENSE_TRAVEL: (
        "Car, van and travel expenses",
        "SA103S Box 12", "SA103F Box 20",
        "BIM45000", False,
    ),
    _cats.SE_EXPENSE_PREMISES: (
        "Rent, rates, power and insurance",
        "SA103S Box 14", "SA103F Box 21",
        "BIM46810", False,
    ),
    _cats.SE_EXPENSE_REPAIRS: (
        "Repairs and renewals",
        "SA103S Box 15", "SA103F Box 22",
        "BIM46900", False,
    ),
    _cats.SE_EXPENSE_ADMIN: (
        "Phone, stationery and other office costs",
        "SA103S Box 18", "SA103F Box 23",
        "BIM47800", False,
    ),
    _cats.SE_EXPENSE_ADVERTISING: (
        "Advertising and business entertainment",
        "SA103S Box 19", "SA103F Box 24",
        "BIM42550", False,
    ),
    _cats.SE_EXPENSE_ENTERTAINMENT: (
        "Business entertainment (added back — disallowed)",
        "SA103S Box 19", "SA103F Box 24",
        "BIM45045", False,
    ),
    _cats.SE_EXPENSE_INTEREST: (
        "Interest on bank and other loans",
        "SA103S Box 17", "SA103F Box 25",
        "BIM45650", False,
    ),
    _cats.SE_EXPENSE_FINANCIAL: (
        "Bank, credit-card and other financial charges",
        "SA103S Box 17", "SA103F Box 26",
        "BIM45650", False,
    ),
    _cats.SE_EXPENSE_BAD_DEBT: (
        "Irrecoverable debts written off",
        "SA103S Box 19", "SA103F Box 27",
        "BIM42700", False,
    ),
    _cats.SE_EXPENSE_PROFESSIONAL: (
        "Accountancy, legal and other professional fees",
        "SA103S Box 16", "SA103F Box 28",
        "BIM46400", False,
    ),
    _cats.SE_EXPENSE_DEPRECIATION: (
        "Depreciation and loss/profit on sale of assets (added back)",
        "SA103S Box 19", "SA103F Box 29",
        "BIM46900", False,
    ),
    _cats.SE_EXPENSE_OTHER: (
        "Other allowable business expenses",
        "SA103S Box 19", "SA103F Box 30",
        "BIM47800", False,
    ),
}

_PROP_META: dict[str, tuple[str, str, str, str, bool]] = {
    _cats.PROP_INCOME_RENT: (
        "Total rents and other receipts",
        "SA105 Box 20", "SA105 Box 20",
        "PIM1051", True,
    ),
    _cats.PROP_INCOME_PREMIUMS: (
        "Premiums on lease grants",
        "SA105 Box 21", "SA105 Box 21",
        "PIM1200", True,
    ),
    _cats.PROP_INCOME_OTHER: (
        "Other property income",
        "SA105 Box 20", "SA105 Box 20",
        "PIM1054", True,
    ),
    _cats.PROP_EXPENSE_PREMISES: (
        "Rent, rates, insurance, ground rents etc.",
        "SA105 Box 24", "SA105 Box 24",
        "PIM2020", False,
    ),
    _cats.PROP_EXPENSE_REPAIRS: (
        "Property repairs and maintenance",
        "SA105 Box 25", "SA105 Box 25",
        "PIM2030", False,
    ),
    _cats.PROP_EXPENSE_FINANCIAL: (
        "Loan interest and other financial costs",
        "SA105 Box 26", "SA105 Box 26",
        "PIM2054", False,
    ),
    _cats.PROP_EXPENSE_RESIDENTIAL_FINANCIAL: (
        "Residential property finance costs (restricted relief)",
        "SA105 Box 44", "SA105 Box 44",
        "PIM2054", False,
    ),
    _cats.PROP_EXPENSE_PROFESSIONAL: (
        "Legal, management and other professional fees",
        "SA105 Box 27", "SA105 Box 27",
        "PIM2040", False,
    ),
    _cats.PROP_EXPENSE_SERVICES: (
        "Costs of services provided, including wages",
        "SA105 Box 28", "SA105 Box 28",
        "PIM2042", False,
    ),
    _cats.PROP_EXPENSE_TRAVEL: (
        "Travel costs",
        "SA105 Box 24", "SA105 Box 24",
        "BIM45000", False,
    ),
    _cats.PROP_EXPENSE_OTHER: (
        "Other allowable property expenses",
        "SA105 Box 29", "SA105 Box 29",
        "PIM2068", False,
    ),
}


def category_meta(category: str | None) -> dict:
    """Look up display metadata for an HMRC category code. Returns sensible
    defaults for unknown / uncategorised."""
    if not category:
        return {
            "label": "Uncategorised — review needed",
            "box_short": "—", "box_full": "—",
            "hmrc_ref": "—", "is_income": False,
            "business_type": "—",
        }
    # Audit-summary uses a sentinel "uncategorised_income" bucket for
    # credits that haven't been categorised yet. Surface it clearly.
    if category == "uncategorised_income":  # noqa: HMRC005 internal sentinel
        return {
            "label": "Uncategorised income — review needed",
            "box_short": "—", "box_full": "—",
            "hmrc_ref": "—", "is_income": True,
            "business_type": "—",
        }
    # Same sentinel for expenses.
    if category == "uncategorised":  # noqa: HMRC005 internal sentinel
        return {
            "label": "Uncategorised expenses — review needed",
            "box_short": "—", "box_full": "—",
            "hmrc_ref": "—", "is_income": False,
            "business_type": "—",
        }
    if category in _SE_META:
        label, sshort, sfull, ref, is_inc = _SE_META[category]
        return {
            "label": label, "box_short": sshort, "box_full": sfull,
            "hmrc_ref": ref, "is_income": is_inc, "business_type": "Self-employment",
        }
    if category in _PROP_META:
        label, sshort, sfull, ref, is_inc = _PROP_META[category]
        return {
            "label": label, "box_short": sshort, "box_full": sfull,
            "hmrc_ref": ref, "is_income": is_inc, "business_type": "UK property",
        }
    # Unknown code — surface so the accountant flags it
    return {
        "label": f"Unrecognised code: {category}",
        "box_short": "—", "box_full": "—",
        "hmrc_ref": "—", "is_income": False, "business_type": "—",
    }


# ---------------------------------------------------------------------------
# Workbook styling
# ---------------------------------------------------------------------------

NAVY = "1B4F72"
LIGHT_NAVY = "D6EAF8"
BAND = "F2F4F4"
WARN = "FCF3CF"
GREEN = "1A7A2E"
RED = "C0392B"

H1 = Font(name="Calibri", bold=True, size=16, color=NAVY)
H2 = Font(name="Calibri", bold=True, size=13, color=NAVY)
HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color=NAVY, end_color=NAVY)
SUBHEADER = Font(name="Calibri", bold=True, size=11, color=NAVY)
SUBHEADER_FILL = PatternFill("solid", start_color=LIGHT_NAVY, end_color=LIGHT_NAVY)
BAND_FILL = PatternFill("solid", start_color=BAND, end_color=BAND)
WARN_FILL = PatternFill("solid", start_color=WARN, end_color=WARN)
MUTED = Font(name="Calibri", size=10, color="566573")
BODY = Font(name="Calibri", size=10)
BODY_BOLD = Font(name="Calibri", bold=True, size=10)
INCOME_FONT = Font(name="Calibri", size=10, color=GREEN)
EXPENSE_FONT = Font(name="Calibri", size=10, color=RED)
THIN = Side(style="thin", color="D5D8DC")
BOTTOM_BORDER = Border(bottom=THIN)

GBP_FMT = '"£"#,##0.00;[Red]-"£"#,##0.00'
PCT_FMT = "0%"


def _autosize(ws, min_w: int = 10, max_w: int = 55) -> None:
    """Approximate Excel auto-fit — openpyxl can't do real auto-fit, so we
    measure the longest value in each column. Good enough for accountants."""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        longest = 0
        for c in col_cells:
            if c.value is None:
                continue
            longest = max(longest, len(str(c.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            max(min_w, min(longest + 2, max_w))
        )


def _band_rows(ws, start_row: int, end_row: int, cols: int) -> None:
    """Banded rows: every other row gets a light fill — accountant scan-helper."""
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, cols + 1):
                ws.cell(row=r, column=c).fill = BAND_FILL


def _set_print_layout(ws, *, portrait: bool = True,
                      fit_to_width: bool = True,
                      title_rows: int = 4) -> None:
    """Configure print-ready layout: A4 portrait, fit-to-width, sensible
    margins, repeat header rows on each printed page. Practice firms
    still print packs; making them look right when you File → Print is
    the difference between "professional" and "tech-bro Excel dump"."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = (
        ws.ORIENTATION_PORTRAIT if portrait else ws.ORIENTATION_LANDSCAPE
    )
    if fit_to_width:
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.5, bottom=0.5,
        header=0.3, footer=0.3,
    )
    ws.print_options.horizontalCentered = True
    if title_rows > 0:
        ws.print_title_rows = f"1:{title_rows}"


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _write_cover(ws, *, user_email: str, period_label: str,
                 client_name: str | None, summary: dict, generated_at: str,
                 period_is_filtered: bool) -> None:
    """A4-print-ready cover page. Sections, in order:
       • Practice header band
       • Client / period block (left) + key totals block (right)
       • Where to start
       • Issues snapshot (count of action items)
       • Sign-off block (Reviewed by / Date / Signature)
       • PI-insurance-friendly disclaimer footer
    """
    ws.title = "Cover"
    totals = summary.get("totals", {})

    # --- Header band ---
    ws.merge_cells("A1:D1")
    ws["A1"] = "BankScan AI — Accountant Pack"
    ws["A1"].font = Font(name="Calibri", bold=True, size=20, color="FFFFFF")
    ws["A1"].fill = HEADER_FILL
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        f"Prepared for professional review · "
        f"{'FILTERED' if period_is_filtered else 'ALL TIME'} · "
        f"Generated {generated_at}"
    )
    ws["A2"].font = Font(name="Calibri", size=10, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", start_color="2E86C1", end_color="2E86C1")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # --- Client / Period block (left) ---
    r = 4
    ws.cell(row=r, column=1, value="Client").font = H2
    r += 1
    block_thin = Border(
        left=THIN, right=THIN, top=THIN, bottom=THIN,
    )
    for label, value in [
        ("Client / business name", client_name or user_email),
        ("Account email", user_email),
        ("Period covered", period_label),
        ("Pack scope",
         "Filtered to period" if period_is_filtered else "All transactions (no date filter)"),
    ]:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BODY_BOLD
        c1.border = block_thin
        c1.fill = BAND_FILL
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = BODY
        c2.border = block_thin
        r += 1

    # --- Key totals block (right) ---
    r_totals = 4
    ws.cell(row=r_totals, column=3, value="Key totals").font = H2
    r_totals += 1
    for label, value in [
        ("Income (gross)", float(totals.get("income", 0) or 0)),
        ("Expenses (gross)", float(totals.get("expenses", 0) or 0)),
        ("Net profit", float(
            (totals.get("income", 0) or 0) - (totals.get("expenses", 0) or 0))),
        ("VAT recorded", float(totals.get("vat_total", 0) or 0)),
    ]:
        c1 = ws.cell(row=r_totals, column=3, value=label)
        c1.font = BODY_BOLD
        c1.border = block_thin
        c1.fill = BAND_FILL
        c2 = ws.cell(row=r_totals, column=4, value=value)
        c2.number_format = GBP_FMT
        c2.font = BODY_BOLD
        c2.border = block_thin
        c2.alignment = Alignment(horizontal="right")
        r_totals += 1

    # --- Transaction counts row (full width) ---
    r = max(r, r_totals) + 1
    ws.cell(row=r, column=1, value="Coverage").font = H2
    r += 1
    coverage_rows = [
        ("Transactions in pack", str(totals.get("transactions_total", 0))),
        ("Backed by a receipt",
         f"{totals.get('transactions_matched', 0)} "
         f"({totals.get('audit_ready_pct', 0)}%)"),
        ("Missing a receipt", str(totals.get("transactions_missing", 0))),
        ("Excluded (personal / owner draws)",
         str(totals.get("transactions_excluded", 0))),
    ]
    for label, value in coverage_rows:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BODY_BOLD
        c1.border = block_thin
        c1.fill = BAND_FILL
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = BODY
        c2.border = block_thin
        r += 1

    # --- Where to start ---
    r += 1
    ws.cell(row=r, column=1, value="Where to start").font = H2
    r += 1
    instructions = [
        "1. Open 'Action Items' — every transaction needing your decision in one place.",
        "2. 'Tax Return Boxes' — totals mapped to SA103S / SA103F / SA105 box numbers.",
        "3. 'Trial Balance' — debit / credit by nominal code; importable into",
        "   CCH / TaxCalc / IRIS / Sage.",
        "4. 'Missing Receipts' is the chase list before submission.",
        "5. summary.html in the ZIP root is the Audit Confidence Certificate;",
        "   open in any browser and Print to PDF for the client file.",
        "6. receipts/ in the ZIP root is grouped by HMRC category for easy review.",
    ]
    for line in instructions:
        ws.cell(row=r, column=1, value=line).font = BODY
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    # --- Sign-off block ---
    r += 1
    ws.cell(row=r, column=1, value="Sign-off").font = H2
    r += 1
    for label in ["Reviewed by", "Position / role", "Date", "Signature"]:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = BODY_BOLD
        c1.border = block_thin
        c1.fill = BAND_FILL
        # Leave columns 2-4 blank for the accountant to fill in by hand
        for col in (2, 3, 4):
            ws.cell(row=r, column=col, value="").border = block_thin
        ws.row_dimensions[r].height = 22
        r += 1

    # --- Disclaimer footer ---
    r += 1
    disclaimer = (
        "AI-assisted draft for professional review. BankScan AI does NOT "
        "provide accounting, tax, or legal advice. Figures require sign-off "
        "by a qualified accountant or tax practitioner. The accompanying "
        "receipts and manifest.json provide a tamper-evident audit trail."
    )
    ws.cell(row=r, column=1, value=disclaimer).font = MUTED
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=4)
    ws.row_dimensions[r].height = 16
    ws.row_dimensions[r + 1].height = 16
    ws.row_dimensions[r + 2].height = 16

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    # Print as one page, A4 portrait.
    _set_print_layout(ws, portrait=True, fit_to_width=True, title_rows=2)


def _detect_issues(tx: dict, linked_receipts: list[dict]) -> list[tuple[str, str, str]]:
    """For one transaction, return a list of (severity, issue, action)
    tuples describing every reason this row needs an accountant's eye.

    Severities: "high" (P0 audit risk), "med" (worth checking), "low" (FYI).
    """
    issues: list[tuple[str, str, str]] = []
    amount = float(tx.get("amount") or 0)
    abs_amount = abs(amount)
    cat = tx.get("hmrc_category")
    conf = tx.get("hmrc_category_confidence")
    status = tx.get("receipt_status") or "missing"

    # P0 — uncategorised at all
    if not cat:
        issues.append((
            "high",
            "No HMRC category assigned",
            "Categorise in BankScan, then re-export.",
        ))

    # P0 — confidence below 50%
    elif conf is not None and conf < 50:
        issues.append((
            "high",
            f"Low confidence ({conf}%) on category {cat}",
            "Verify the category matches the transaction.",
        ))

    # P1 — missing receipt on a large expense
    if status not in ("matched", "excluded") and amount < 0 and abs_amount >= 100:
        issues.append((
            "med",
            f"Expense £{abs_amount:.2f} has no receipt",
            "Chase the receipt or mark as personal/excluded.",
        ))

    # P1 — sign mismatch: matched receipt on an income row
    if linked_receipts and amount > 0:
        issues.append((
            "high",
            "Income credit linked to an expense receipt",
            "Unmatch the receipt — receipts only belong on expenses.",
        ))

    # P2 — VAT > 25% of gross (UK standard rate is 20%, so anything over
    # 25% is almost always a parser slip)
    vat = tx.get("vat_amount")
    if vat is not None and abs_amount > 0 and (float(vat) / abs_amount) > 0.25:
        issues.append((
            "med",
            f"VAT £{float(vat):.2f} is >25% of gross £{abs_amount:.2f}",
            "Verify the VAT line on the receipt.",
        ))

    return issues


def _write_action_items(ws, txs: list[dict],
                        links_by_tx: dict[int, list[dict]],
                        receipts_by_id: dict[int, dict]) -> None:
    """The very first data sheet the accountant opens after the Cover.
    One row per (transaction, issue) pair so a single transaction with
    two problems shows on two rows."""
    ws.title = "Action Items"

    headers = [
        "Severity", "Tx ID", "Date", "Description", "Amount (GBP)",
        "HMRC category", "Issue", "Suggested action",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    sev_fill = {
        "high": PatternFill("solid", start_color="F5B7B1", end_color="F5B7B1"),
        "med":  WARN_FILL,
        "low":  PatternFill("solid", start_color="D5DBDB", end_color="D5DBDB"),
    }
    sev_rank = {"high": 0, "med": 1, "low": 2}

    rows: list[dict] = []
    for tx in txs:
        linked_full = [
            receipts_by_id.get(l["receipt_id"]) or {}
            for l in links_by_tx.get(tx["id"], [])
        ]
        for severity, issue, action in _detect_issues(tx, linked_full):
            rows.append({
                "severity": severity, "tx": tx, "issue": issue, "action": action,
            })

    # Highest severity first, then by amount magnitude
    rows.sort(key=lambda r: (
        sev_rank.get(r["severity"], 9),
        -abs(float(r["tx"].get("amount") or 0)),
    ))

    if not rows:
        # No issues — give the accountant a clear "all green" message
        ws.cell(row=2, column=1,
                value="✓ No issues detected. Pack is ready for review.").font = INCOME_FONT
        ws.merge_cells("A2:H2")
    else:
        for i, row in enumerate(rows, start=2):
            tx = row["tx"]
            ws.cell(row=i, column=1, value=row["severity"].upper()).font = BODY_BOLD
            ws.cell(row=i, column=1).fill = sev_fill[row["severity"]]
            ws.cell(row=i, column=2, value=tx["id"]).font = BODY
            ws.cell(row=i, column=3, value=tx.get("date_iso") or "").font = BODY
            ws.cell(row=i, column=4, value=tx.get("description") or "").font = BODY
            amt_cell = ws.cell(row=i, column=5, value=float(tx.get("amount") or 0))
            amt_cell.number_format = GBP_FMT
            amt_cell.font = INCOME_FONT if float(tx.get("amount") or 0) > 0 else EXPENSE_FONT
            ws.cell(row=i, column=6, value=tx.get("hmrc_category") or "—").font = BODY
            ws.cell(row=i, column=7, value=row["issue"]).font = BODY
            ws.cell(row=i, column=8, value=row["action"]).font = BODY

        ws.auto_filter.ref = ws.dimensions

    ws.freeze_panes = "A2"
    _autosize(ws, max_w=50)


def _write_trial_balance(ws, summary: dict) -> None:
    """The format every UK practice tool imports: nominal code, description,
    debit, credit, net. Built straight from the audit-summary buckets so it
    reconciles to the Tax Return Boxes sheet pound-for-pound."""
    ws.title = "Trial Balance"
    headers = ["Nominal code", "Description", "Debit (GBP)", "Credit (GBP)", "Net (GBP)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    cats = summary.get("categories", [])
    r = 2
    total_debit = 0.0
    total_credit = 0.0
    for cat in sorted(
        cats, key=lambda c: (not c.get("is_income"), c.get("category") or "")
    ):
        code = cat.get("category") or "uncategorised"
        meta = category_meta(code)
        amt = float(cat.get("total_gross_gbp") or 0)
        if cat.get("is_income"):
            debit, credit = 0.0, amt
            total_credit += amt
        else:
            debit, credit = amt, 0.0
            total_debit += amt
        ws.cell(row=r, column=1, value=code).font = BODY
        ws.cell(row=r, column=2, value=meta["label"]).font = BODY
        if debit:
            c = ws.cell(row=r, column=3, value=debit)
            c.number_format = GBP_FMT
            c.font = EXPENSE_FONT
        if credit:
            c = ws.cell(row=r, column=4, value=credit)
            c.number_format = GBP_FMT
            c.font = INCOME_FONT
        net_cell = ws.cell(row=r, column=5, value=(credit - debit))
        net_cell.number_format = GBP_FMT
        net_cell.font = BODY
        r += 1

    # Totals row — must reconcile to the totals on the Cover sheet
    r += 1
    ws.cell(row=r, column=2, value="TOTALS").font = BODY_BOLD
    d_total = ws.cell(row=r, column=3, value=total_debit)
    d_total.number_format = GBP_FMT
    d_total.font = BODY_BOLD
    c_total = ws.cell(row=r, column=4, value=total_credit)
    c_total.number_format = GBP_FMT
    c_total.font = BODY_BOLD
    net_total = ws.cell(row=r, column=5, value=(total_credit - total_debit))
    net_total.number_format = GBP_FMT
    net_total.font = BODY_BOLD

    # Reconciliation note — accountants like a "this should equal that" line
    r += 2
    ws.cell(row=r, column=2, value=(
        "The Net (credit − debit) column should match the Net profit on "
        "the Cover sheet."
    )).font = MUTED
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{r-2}"
    _autosize(ws)
    _set_print_layout(ws, portrait=True, fit_to_width=True, title_rows=1)


def _write_boxes(ws, summary: dict) -> None:
    """The headline sheet for the accountant: each category mapped to its
    SA box number with the ready-to-transcribe total."""
    ws.title = "Tax Return Boxes"

    ws["A1"] = "Tax Return Boxes — ready to transcribe"
    ws["A1"].font = H1
    ws.merge_cells("A1:G1")
    ws["A2"] = (
        "Each line is the total to enter on the corresponding SA box. "
        "Short / full pages shown side-by-side."
    )
    ws["A2"].font = MUTED
    ws.merge_cells("A2:G2")

    headers = [
        "HMRC code", "Description", "SA short", "SA full",
        "Total (gross)", "VAT", "Receipts backing",
    ]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    cats = summary.get("categories", [])
    # Split: income first, expenses second — same order as SA pages
    income_cats = [c for c in cats if c.get("is_income")]
    expense_cats = [c for c in cats if not c.get("is_income")]

    r = header_row + 1
    if income_cats:
        ws.cell(row=r, column=1, value="INCOME").font = SUBHEADER
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = SUBHEADER_FILL
        r += 1
        income_start = r
        for cat in income_cats:
            meta = category_meta(cat["category"])
            ws.cell(row=r, column=1, value=cat["category"]).font = BODY
            ws.cell(row=r, column=2, value=meta["label"]).font = BODY
            ws.cell(row=r, column=3, value=meta["box_short"]).font = BODY
            ws.cell(row=r, column=4, value=meta["box_full"]).font = BODY
            amt_cell = ws.cell(row=r, column=5, value=cat.get("total_gross_gbp") or 0)
            amt_cell.font = INCOME_FONT
            amt_cell.number_format = GBP_FMT
            vat_cell = ws.cell(row=r, column=6, value=cat.get("total_vat_gbp") or 0)
            vat_cell.number_format = GBP_FMT
            vat_cell.font = BODY
            ws.cell(row=r, column=7,
                    value=f"{cat.get('matched_count', 0)}/{cat.get('transaction_count', 0)}"
                    ).font = BODY
            r += 1
        _band_rows(ws, income_start, r - 1, 7)

    if expense_cats:
        ws.cell(row=r, column=1, value="EXPENSES").font = SUBHEADER
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = SUBHEADER_FILL
        r += 1
        exp_start = r
        for cat in expense_cats:
            meta = category_meta(cat["category"])
            ws.cell(row=r, column=1, value=cat["category"]).font = BODY
            ws.cell(row=r, column=2, value=meta["label"]).font = BODY
            ws.cell(row=r, column=3, value=meta["box_short"]).font = BODY
            ws.cell(row=r, column=4, value=meta["box_full"]).font = BODY
            amt_cell = ws.cell(row=r, column=5, value=cat.get("total_gross_gbp") or 0)
            amt_cell.font = EXPENSE_FONT
            amt_cell.number_format = GBP_FMT
            vat_cell = ws.cell(row=r, column=6, value=cat.get("total_vat_gbp") or 0)
            vat_cell.number_format = GBP_FMT
            vat_cell.font = BODY
            ws.cell(row=r, column=7,
                    value=f"{cat.get('matched_count', 0)}/{cat.get('transaction_count', 0)}"
                    ).font = BODY
            # Flag categories with no receipt backing — accountant scan-cue
            if cat.get("transaction_count") and cat.get("matched_count") == 0:
                for col in range(1, 8):
                    ws.cell(row=r, column=col).fill = WARN_FILL
            r += 1
        _band_rows(ws, exp_start, r - 1, 7)

    # Totals row
    r += 1
    ws.cell(row=r, column=2, value="TOTAL INCOME").font = BODY_BOLD
    inc_cell = ws.cell(row=r, column=5, value=summary.get("totals", {}).get("income", 0))
    inc_cell.font = BODY_BOLD
    inc_cell.number_format = GBP_FMT
    r += 1
    ws.cell(row=r, column=2, value="TOTAL EXPENSES").font = BODY_BOLD
    exp_cell = ws.cell(row=r, column=5, value=summary.get("totals", {}).get("expenses", 0))
    exp_cell.font = BODY_BOLD
    exp_cell.number_format = GBP_FMT
    r += 1
    ws.cell(row=r, column=2, value="NET PROFIT (income − expenses)").font = BODY_BOLD
    net = (summary.get("totals", {}).get("income", 0) -
           summary.get("totals", {}).get("expenses", 0))
    net_cell = ws.cell(row=r, column=5, value=net)
    net_cell.font = BODY_BOLD
    net_cell.number_format = GBP_FMT

    ws.freeze_panes = "A5"
    _autosize(ws)
    _set_print_layout(ws, portrait=True, fit_to_width=True, title_rows=4)


def _write_transactions(ws, txs: list[dict], links_by_tx: dict[int, list[dict]],
                        receipts_by_id: dict[int, dict]) -> None:
    ws.title = "Transactions"
    headers = [
        "ID", "Date", "Description", "Amount (GBP)", "HMRC category",
        "Category friendly", "SA box (short)",
        "Confidence (%)", "Business %", "Capital", "Status",
        "Receipt(s)", "VAT", "Hash (first 8)",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    for i, tx in enumerate(txs, start=2):
        meta = category_meta(tx.get("hmrc_category"))
        linked = links_by_tx.get(tx["id"], [])
        receipt_summary = "; ".join(
            f"{(receipts_by_id.get(l['receipt_id']) or {}).get('store_name') or 'rc'+str(l['receipt_id'])}"
            f" £{(receipts_by_id.get(l['receipt_id']) or {}).get('total_amount') or 0:.2f}"
            for l in linked
        )
        ws.cell(row=i, column=1, value=tx["id"]).font = BODY
        ws.cell(row=i, column=2, value=tx.get("date_iso") or "").font = BODY
        ws.cell(row=i, column=3, value=tx.get("description") or "").font = BODY
        amt_cell = ws.cell(row=i, column=4, value=float(tx.get("amount") or 0))
        amt_cell.font = INCOME_FONT if float(tx.get("amount") or 0) > 0 else EXPENSE_FONT
        amt_cell.number_format = GBP_FMT
        ws.cell(row=i, column=5, value=tx.get("hmrc_category") or "").font = BODY
        ws.cell(row=i, column=6, value=meta["label"]).font = BODY
        ws.cell(row=i, column=7, value=meta["box_short"]).font = BODY
        ws.cell(row=i, column=8,
                value=tx.get("hmrc_category_confidence") or "").font = BODY
        ws.cell(row=i, column=9, value=tx.get("business_pct") or 100).font = BODY
        ws.cell(row=i, column=10, value="Yes" if tx.get("is_capital") else "").font = BODY
        ws.cell(row=i, column=11, value=tx.get("receipt_status") or "missing").font = BODY
        ws.cell(row=i, column=12, value=receipt_summary).font = BODY
        vat_v = tx.get("vat_amount")
        if vat_v is not None:
            vat_cell = ws.cell(row=i, column=13, value=float(vat_v))
            vat_cell.number_format = GBP_FMT
            vat_cell.font = BODY
        ws.cell(row=i, column=14,
                value=(tx.get("content_hash") or "")[:8]).font = MUTED

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, max_w=45)


def _write_missing(ws, txs: list[dict]) -> None:
    ws.title = "Missing Receipts"
    headers = [
        "ID", "Date", "Description", "Amount (GBP)",
        "HMRC category", "Category friendly", "Status",
        "Exclusion reason",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    r = 2
    for tx in txs:
        if tx.get("receipt_status") in ("matched", "excluded"):
            continue
        meta = category_meta(tx.get("hmrc_category"))
        ws.cell(row=r, column=1, value=tx["id"]).font = BODY
        ws.cell(row=r, column=2, value=tx.get("date_iso") or "").font = BODY
        ws.cell(row=r, column=3, value=tx.get("description") or "").font = BODY
        amt_cell = ws.cell(row=r, column=4, value=float(tx.get("amount") or 0))
        amt_cell.font = EXPENSE_FONT
        amt_cell.number_format = GBP_FMT
        ws.cell(row=r, column=5, value=tx.get("hmrc_category") or "").font = BODY
        ws.cell(row=r, column=6, value=meta["label"]).font = BODY
        ws.cell(row=r, column=7, value=tx.get("receipt_status") or "missing").font = BODY
        ws.cell(row=r, column=8, value=tx.get("exclusion_reason") or "").font = BODY
        r += 1

    ws.freeze_panes = "A2"
    if r > 2:
        ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _write_receipt_inventory(ws, receipts: list[dict],
                             links_by_rc: dict[int, list[dict]]) -> None:
    ws.title = "Receipt Inventory"
    headers = [
        "Receipt ID", "Date", "Store", "Total (GBP)", "VAT",
        "Status", "Linked to tx ID(s)", "Source file",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    for i, r in enumerate(receipts, start=2):
        linked_tx = links_by_rc.get(r["id"], [])
        linked_txt = ", ".join(str(l["transaction_id"]) for l in linked_tx)
        status = "matched" if linked_tx else "orphan"
        ws.cell(row=i, column=1, value=r["id"]).font = BODY
        ws.cell(row=i, column=2, value=r.get("date_iso") or "").font = BODY
        ws.cell(row=i, column=3, value=r.get("store_name") or "Unknown").font = BODY
        amt = r.get("total_amount")
        if amt is not None:
            amt_cell = ws.cell(row=i, column=4, value=float(amt))
            amt_cell.number_format = GBP_FMT
            amt_cell.font = BODY
        vat = r.get("tax_amount")
        if vat is not None:
            vat_cell = ws.cell(row=i, column=5, value=float(vat))
            vat_cell.number_format = GBP_FMT
            vat_cell.font = BODY
        ws.cell(row=i, column=6, value=status).font = BODY
        ws.cell(row=i, column=7, value=linked_txt).font = BODY
        ws.cell(row=i, column=8, value=r.get("source_filename") or "").font = MUTED

    ws.freeze_panes = "A2"
    if len(receipts) > 0:
        ws.auto_filter.ref = ws.dimensions
    _autosize(ws, max_w=45)


def _write_vat(ws, txs: list[dict]) -> None:
    ws.title = "VAT Register"
    headers = [
        "ID", "Date", "Description", "Net (calculated)",
        "VAT", "Gross (amount)", "HMRC category",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    r = 2
    for tx in txs:
        if not tx.get("vat_amount"):
            continue
        gross = abs(float(tx.get("amount") or 0))
        vat = float(tx["vat_amount"])
        net = max(gross - vat, 0)
        ws.cell(row=r, column=1, value=tx["id"]).font = BODY
        ws.cell(row=r, column=2, value=tx.get("date_iso") or "").font = BODY
        ws.cell(row=r, column=3, value=tx.get("description") or "").font = BODY
        for col_idx, val in [(4, net), (5, vat), (6, gross)]:
            c = ws.cell(row=r, column=col_idx, value=val)
            c.number_format = GBP_FMT
            c.font = BODY
        ws.cell(row=r, column=7, value=tx.get("hmrc_category") or "").font = BODY
        r += 1

    ws.freeze_panes = "A2"
    if r > 2:
        ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def _write_reasoning(ws, txs: list[dict]) -> None:
    ws.title = "Reasoning Log"
    headers = [
        "Tx ID", "Date", "Description", "HMRC code",
        "Category friendly", "Confidence (%)", "AI reasoning",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER
        cell.fill = HEADER_FILL

    for i, tx in enumerate(txs, start=2):
        meta = category_meta(tx.get("hmrc_category"))
        ws.cell(row=i, column=1, value=tx["id"]).font = BODY
        ws.cell(row=i, column=2, value=tx.get("date_iso") or "").font = BODY
        ws.cell(row=i, column=3, value=tx.get("description") or "").font = BODY
        ws.cell(row=i, column=4, value=tx.get("hmrc_category") or "").font = BODY
        ws.cell(row=i, column=5, value=meta["label"]).font = BODY
        ws.cell(row=i, column=6,
                value=tx.get("hmrc_category_confidence") or "").font = BODY
        ws.cell(row=i, column=7,
                value=tx.get("hmrc_category_reason") or "").font = BODY

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws, max_w=70)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def build_accountant_workbook(
    *,
    user_email: str,
    period_label: str,
    client_name: str | None,
    summary: dict,
    txs: list[dict],
    receipts: list[dict],
    links_by_tx: dict[int, list[dict]],
    links_by_rc: dict[int, list[dict]],
    generated_at: str | None = None,
    period_is_filtered: bool = False,
) -> bytes:
    """Build the multi-sheet Accountant Pack workbook and return bytes.

    Sheet order is deliberate: ALL data the accountant must act on is
    front-loaded.
      1. Cover           — what this is, who it's for, sign-off
      2. Action Items    — every row needing a decision, severity-sorted
      3. Tax Return Boxes — totals ready to transcribe onto SA103/SA105
      4. Trial Balance   — debit/credit per nominal for software import
      5. Transactions    — the full ledger (filterable)
      6. Missing Receipts
      7. Receipt Inventory
      8. VAT Register
      9. Reasoning Log
    """
    if generated_at is None:
        generated_at = _dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")

    wb = Workbook()
    receipts_by_id = {r["id"]: r for r in receipts}

    _write_cover(
        wb.active,
        user_email=user_email, period_label=period_label,
        client_name=client_name, summary=summary, generated_at=generated_at,
        period_is_filtered=period_is_filtered,
    )
    _write_action_items(wb.create_sheet(), txs, links_by_tx, receipts_by_id)
    _write_boxes(wb.create_sheet(), summary=summary)
    _write_trial_balance(wb.create_sheet(), summary=summary)
    _write_transactions(wb.create_sheet(), txs, links_by_tx, receipts_by_id)
    _write_missing(wb.create_sheet(), txs)
    _write_receipt_inventory(wb.create_sheet(), receipts, links_by_rc)
    _write_vat(wb.create_sheet(), txs)
    _write_reasoning(wb.create_sheet(), txs)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
